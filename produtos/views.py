from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from .models import Estoque
from django.core.paginator import Paginator
from datetime import date, timedelta

# View para relatório completo do estoque
@login_required
def relatorio_completo(request):
    estoque = Estoque.objects.select_related('produto', 'local').order_by('-data_armazenado')
    today = date.today()
    hoje_mais_30 = today + timedelta(days=30)
    return render(request, 'produtos/relatorio_completo.html', {
        'estoque': estoque,
        'today': today,
        'hoje_mais_30': hoje_mais_30
    })

# View para relatório paginado do estoque
@login_required
def relatorio_estoque(request):
    dados = Estoque.objects.select_related('produto', 'local').order_by('-data_armazenado')
    today = date.today()
    hoje_mais_30 = today + timedelta(days=30)
    paginator = Paginator(dados, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'produtos/relatorios.html', {
        'page_obj': page_obj,
        'today': today,
        'hoje_mais_30': hoje_mais_30
    })
    today = date.today()
    hoje_mais_30 = today + timedelta(days=30)
    return render(request, 'produtos/relatorio_completo.html', {
        'estoque': estoque,
        'today': today,
        'hoje_mais_30': hoje_mais_30
    })
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from .models import Produto, Armazenamento, Estoque, Lote, HistoricoMovimentacao, Alerta
from django.utils import timezone
import csv
import io
from django.http import HttpResponse, JsonResponse
from django.core.paginator import Paginator
from django.db.models import Q, Count
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from .forms import ProdutoForm, ArmazenamentoForm
from datetime import date, timedelta
import json

@login_required
def dashboard(request):
    """Dashboard principal com métricas e alertas"""
    from datetime import date, timedelta
    
    # Estatísticas básicas
    total_produtos = Produto.objects.count()
    produtos_em_estoque = Estoque.objects.values('produto').distinct().count()
    total_enderecos = Armazenamento.objects.count()
    enderecos_com_estoque = Estoque.objects.values('local').distinct().count()
    
    # Taxa de ocupação
    taxa_ocupacao = round((enderecos_com_estoque / total_enderecos * 100), 1) if total_enderecos > 0 else 0
    
    # Produtos por status de validade
    hoje = date.today()
    proximos_30_dias = hoje + timedelta(days=30)
    
    produtos_vencidos = 0
    produtos_vencendo = 0
    produtos_validos = 0
    
    for produto in Produto.objects.all():
        proxima_validade = produto.proxima_validade()
        if proxima_validade:
            if proxima_validade < hoje:
                produtos_vencidos += 1
            elif proxima_validade <= proximos_30_dias:
                produtos_vencendo += 1
            else:
                produtos_validos += 1
    
    # Alertas ativos
    try:
        alertas_ativos = Alerta.objects.filter(status='ativo').order_by('-data_criacao')[:5]
    except:
        alertas_ativos = []
    
    # Últimas movimentações
    try:
        ultimas_movimentacoes = HistoricoMovimentacao.objects.select_related('produto').order_by('-data_operacao')[:10]
    except:
        ultimas_movimentacoes = []
    
    context = {
        'total_produtos': total_produtos,
        'produtos_em_estoque': produtos_em_estoque,
        'total_enderecos': total_enderecos,
        'taxa_ocupacao': taxa_ocupacao,
        'produtos_vencidos': produtos_vencidos,
        'produtos_vencendo': produtos_vencendo,
        'produtos_validos': produtos_validos,
        'alertas_ativos': alertas_ativos,
        'ultimas_movimentacoes': ultimas_movimentacoes,
    }
    
    return render(request, 'produtos/dashboard.html', context)

def buscar_produto(request):
    if request.method == 'POST':
        codigo = request.POST.get('codigo')
        try:
            produto = Produto.objects.get(codigo=codigo)
            return redirect('perguntar_armazenar',produto_id=produto.id)
        except Produto.DoesNotExist:
            messages.warning(request,'Produto não encontrado. Redirecionando para cadastro.')
            return redirect('cadastrar_produto')
    return render(request,'produtos/buscar_produto.html')

def perguntar_armazenar(request,produto_id):
    produto = get_object_or_404(Produto,id=produto_id)
    if request.method == 'POST':
        escolha = request.POST.get('armazenar')
        if escolha == 'sim':
            return redirect('armazenar_produto',produto_id=produto.id)
        else:
            messages.info(request,'operaçao cancelada')
            return redirect('buscar_produto')
    return render(request,'produtos/perguntar_armazenar.html',{'produto':produto})
 
def armazenar_produto(request, produto_id=None):
    produto = None
    if produto_id:
        produto = get_object_or_404(Produto, id=produto_id)
    enderecos_disponiveis = list(Armazenamento.objects.filter(livre=True))
    # Se produto já está armazenado, inclui o endereço atual na lista
    if produto:
        estoque_atual = Estoque.objects.filter(produto=produto).first()
        if estoque_atual and estoque_atual.local not in enderecos_disponiveis:
            enderecos_disponiveis.append(estoque_atual.local)
    if request.method == 'POST':
        validade = request.POST.get('validade')
        data_armazenado = request.POST.get('data_armazenado')
        endereco_id = request.POST.get('endereco_id')
        numero_lote = request.POST.get('numero_lote', '')
        quantidade = request.POST.get('quantidade', 1)

        if not produto:
            # Criar novo produto
            codigo = request.POST.get('codigo')
            nome = request.POST.get('nome')
            peso = request.POST.get('peso')
            produto = Produto.objects.create(
                codigo=codigo,
                nome=nome,
                peso=peso
            )

        local = get_object_or_404(Armazenamento, id=endereco_id)

        # Adiciona novo lote
        lote = Lote.objects.create(
            produto=produto,
            validade=validade,
            numero_lote=numero_lote,
            quantidade=quantidade
        )

        # Atualiza ou cria registro de estoque
        estoque_existente = Estoque.objects.filter(produto=produto, local=local).first()
        if estoque_existente:
            # Atualiza data_armazenado para o mais recente, se informado
            if data_armazenado:
                estoque_existente.data_armazenado = data_armazenado
                estoque_existente.save()
            else:
                # Se não vier data, mantém a atual
                pass
        else:
            Estoque.objects.create(produto=produto, local=local, data_armazenado=data_armazenado)
            local.livre = False
            local.save()
        # Criar histórico de movimentação
        HistoricoMovimentacao.objects.create(
            produto=produto,
            local_destino=local,
            tipo_operacao='entrada',
            quantidade=quantidade,
            usuario=request.user.username if request.user.is_authenticated else 'Sistema',
            observacoes=f'Produto armazenado - Lote: {numero_lote}'
        )
        
        # Verificar e criar alertas se necessário
        if validade:
            validade_date = timezone.datetime.strptime(validade, '%Y-%m-%d').date()
            dias_para_vencer = (validade_date - date.today()).days
            
            if dias_para_vencer <= 0:
                Alerta.objects.create(
                    tipo='vencido',
                    titulo=f'Produto Vencido: {produto.nome}',
                    mensagem=f'O produto {produto.nome} (código {produto.codigo}) está vencido desde {validade_date.strftime("%d/%m/%Y")}',
                    produto=produto
                )
            elif dias_para_vencer <= 30:
                Alerta.objects.create(
                    tipo='vencimento',
                    titulo=f'Produto Próximo ao Vencimento: {produto.nome}',
                    mensagem=f'O produto {produto.nome} (código {produto.codigo}) vence em {dias_para_vencer} dias ({validade_date.strftime("%d/%m/%Y")})',
                    produto=produto
                )

        messages.success(request, 'Produto armazenado com sucesso!')
        return redirect('painel')

    return render(request, 'produtos/armazenar_produto.html', {
        'produto': produto,
        'codigo': produto.codigo if produto else request.GET.get('codigo', ''),
        'enderecos': enderecos_disponiveis,
        'hoje': timezone.now().date()
    })

@login_required
def relatorio_estoque(request):
    dados = Estoque.objects.select_related('produto', 'local').order_by('-data_armazenado')
    today = date.today()
    hoje_mais_30 = today + timedelta(days=30)
    
    paginator = Paginator(dados, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'produtos/relatorios.html', {
        'page_obj': page_obj,
        'today': today,
        'hoje_mais_30': hoje_mais_30
    })

@login_required
def painel(request):
    """
    Painel principal que exibe o estoque organizando por rua e prédio - OTIMIZADO
    """
    # Importar utils otimizados
    from .utils import EstoqueManager, ValidadeManager
    from django.db.models import Prefetch
    
    # Processar filtros de visualização
    filtro = request.GET.get('filtro', '')
    
    # Processar busca por código
    busca_codigo = request.GET.get('codigo', '')
    resultado_busca = None
    
    if busca_codigo:
        try:
            produto = Produto.objects.get(codigo=busca_codigo)
            resultado_busca = produto
        except Produto.DoesNotExist:
            resultado_busca = 'not_found'
    
    # Usar EstoqueManager para busca otimizada
    if filtro == 'vazios':
        enderecos = EstoqueManager.buscar_enderecos_vazios()[:50]  # Limitar para performance
    elif filtro == 'ocupados':
        enderecos = EstoqueManager.buscar_enderecos_ocupados()
    else:
        enderecos = EstoqueManager.buscar_enderecos_mixto()[:50]  # Limitar para performance
    
    # Organizar por ruas e prédios usando lógica otimizada
    organizacao = {}
    
    # Prefetch dos dados necessários para reduzir queries
    enderecos_com_dados = Armazenamento.objects.filter(
        id__in=[e.id for e in enderecos]
    ).prefetch_related(
        Prefetch('estoque_set', 
            queryset=Estoque.objects.select_related('produto').prefetch_related(
                Prefetch('produto__lotes', 
                    queryset=Lote.objects.order_by('validade'))
            ))
    ).order_by('rua', 'predio', 'nivel')
    
    for endereco in enderecos_com_dados:
        rua = endereco.rua
        predio = endereco.predio
        
        # Inicializar estrutura se não existir
        if rua not in organizacao:
            organizacao[rua] = {}
        if predio not in organizacao[rua]:
            organizacao[rua][predio] = []
        
        # Usar dados já carregados via prefetch
        produtos_estoque = endereco.estoque_set.all()
        
        # Enriquecer dados usando ValidadeManager
        produtos_com_lotes = []
        for estoque in produtos_estoque:
            produto = estoque.produto
            lotes = list(produto.lotes.all())  # Já carregado via prefetch
            
            # Usar ValidadeManager para calcular status
            validade_info = ValidadeManager.calcular_status_validade(lotes)
            
            produto_info = {
                'estoque': estoque,
                'produto': produto,
                'lotes': lotes,
                'proxima_validade': validade_info['proxima_validade'],
                'status_validade': validade_info['status_validade'],
                'dias_para_vencer': validade_info['dias_para_vencer'],
                'total_lotes': len(lotes)
            }
            produtos_com_lotes.append(produto_info)
        
        # Adicionar informações do endereço
        endereco_info = {
            'endereco': endereco,
            'produtos': produtos_com_lotes,
            'tem_produtos': len(produtos_com_lotes) > 0,
            'total_produtos': len(produtos_com_lotes)
        }
        
        organizacao[rua][predio].append(endereco_info)
    
    # Usar EstoqueManager para estatísticas otimizadas
    estatisticas = EstoqueManager.calcular_estatisticas()
    
    context = {
        'organizacao': organizacao,
        'total_enderecos': estatisticas['total_enderecos'],
        'enderecos_com_estoque': estatisticas['enderecos_com_estoque'],
        'enderecos_vazios': estatisticas['enderecos_vazios'],
        'total_produtos': estatisticas['total_produtos'],
        'taxa_ocupacao': estatisticas['taxa_ocupacao'],
        'busca_codigo': busca_codigo,
        'resultado_busca': resultado_busca,
        'produto': resultado_busca if resultado_busca != 'not_found' else None,
        'filtro_ativo': filtro
    }
    
    return render(request, 'produtos/painel.html', context)

def editar_estoque(request, estoque_id):
    """
    View para editar um item específico do estoque:
    - Alterar data de armazenamento
    - Editar observações
    - Excluir do estoque
    """
    estoque = get_object_or_404(Estoque, id=estoque_id)
    produto = estoque.produto
    lotes = produto.lotes.all().order_by('validade')

    if request.method == 'POST':
        if 'excluir_estoque' in request.POST:
            # Liberar endereço
            estoque.local.livre = True
            estoque.local.save()
            
            # Criar histórico de movimentação
            HistoricoMovimentacao.objects.create(
                produto=produto,
                local_origem=estoque.local,
                tipo_operacao='saida',
                quantidade=1,
                usuario=request.user.username if request.user.is_authenticated else 'Sistema',
                observacoes='Produto removido via edição do estoque'
            )
            
            # Remover estoque
            estoque.delete()
            messages.success(request, 'Produto removido do estoque com sucesso!')
            return redirect('painel')
            
        elif 'nova_validade' in request.POST:
            # Adicionar novo lote
            nova_validade = request.POST.get('nova_validade')
            novo_numero_lote = request.POST.get('novo_numero_lote', '')
            nova_quantidade = request.POST.get('nova_quantidade', 1)
            Lote.objects.create(
                produto=produto, 
                validade=nova_validade, 
                numero_lote=novo_numero_lote, 
                quantidade=nova_quantidade
            )
            messages.success(request, 'Novo lote adicionado com sucesso!')
            return redirect('editar_estoque', estoque_id=estoque_id)
            
        elif 'remover_lote_id' in request.POST:
            # Remover lote específico
            remover_id = request.POST.get('remover_lote_id')
            lote_removido = get_object_or_404(Lote, id=remover_id)
            lote_removido.delete()
            messages.success(request, 'Lote removido com sucesso!')
            return redirect('editar_estoque', estoque_id=estoque_id)
            
        else:
            # Atualizar dados do estoque E do produto
            data_armazenado = request.POST.get('data_armazenado')
            observacoes = request.POST.get('observacoes', '')
            
            # Dados do produto
            nome_produto = request.POST.get('nome_produto', '').strip()
            codigo_produto = request.POST.get('codigo_produto', '').strip()
            peso_produto = request.POST.get('peso_produto', '').strip()
            categoria_produto = request.POST.get('categoria_produto', '').strip()
            fornecedor_produto = request.POST.get('fornecedor_produto', '').strip()
            
            # Validar e atualizar produto se houve mudanças
            if nome_produto and nome_produto != produto.nome:
                produto.nome = nome_produto
            if codigo_produto and codigo_produto != produto.codigo:
                # Verificar se o código já existe em outro produto
                if Produto.objects.filter(codigo=codigo_produto).exclude(id=produto.id).exists():
                    messages.error(request, f'Código "{codigo_produto}" já está sendo usado por outro produto!')
                    return redirect('editar_estoque', estoque_id=estoque_id)
                produto.codigo = codigo_produto
            if peso_produto and peso_produto != produto.peso:
                produto.peso = peso_produto
            if categoria_produto and categoria_produto != produto.categoria:
                produto.categoria = categoria_produto
            if fornecedor_produto and fornecedor_produto != produto.fornecedor:
                produto.fornecedor = fornecedor_produto
            
            # Salvar produto
            produto.save()
            
            # Atualizar estoque
            if data_armazenado:
                estoque.data_armazenado = data_armazenado
            estoque.observacoes = observacoes
            estoque.save()
            
            messages.success(request, 'Produto e estoque atualizados com sucesso!')
            return redirect('painel')

    return render(request, 'produtos/editar_estoque.html', {
        'estoque': estoque,
        'produto': produto,
        'lotes': lotes,
    })

def remover_produto(request,estoque_id):
    estoque = get_object_or_404(Estoque, id=estoque_id)
    produto = estoque.produto

    # Criar histórico de movimentação
    HistoricoMovimentacao.objects.create(
        produto=produto,
        local_origem=estoque.local,
        tipo_operacao='saida',
        quantidade=1,  # Assumindo quantidade 1, pode ser ajustado
        usuario=request.user.username if request.user.is_authenticated else 'Sistema',
        observacoes='Produto removido do estoque'
    )

    # Liberar endereço
    estoque.local.livre = True
    estoque.local.save()

    # Remover todos os lotes do produto
    produto.lotes.all().delete()

    # Remover todos os registros de estoque do produto
    Estoque.objects.filter(produto=produto).delete()

    # Mantém o Produto (nome/código) no banco
    return redirect('painel')
    
def cadastrar_produto(request):
    # Capturar parâmetros da URL
    codigo_preenchido = request.GET.get('codigo', '')
    endereco_retorno = request.GET.get('endereco_retorno', '')
    
    if request.method == 'POST':
        form = ProdutoForm(request.POST)
        if form.is_valid():
            produto = form.save()
            messages.success(request, 'Produto cadastrado com sucesso!')
            
            # Se veio de busca por endereço, redirecionar para confirmação de armazenamento
            if endereco_retorno and endereco_retorno.isdigit():
                return redirect('confirmar_armazenamento_endereco', endereco_id=endereco_retorno, produto_id=produto.id)
            else:
                # Permanecer na página de cadastro para novos cadastros
                # Limpar o formulário para um novo produto
                form = ProdutoForm()
                messages.success(request, f'Produto "{produto.nome}" cadastrado! Cadastre outro produto abaixo.')
    else:
        # Pré-preencher o código se fornecido
        initial_data = {}
        if codigo_preenchido:
            initial_data['codigo'] = codigo_preenchido
        form = ProdutoForm(initial=initial_data)
    
    context = {
        'form': form,
        'codigo_preenchido': codigo_preenchido,
        'endereco_retorno': endereco_retorno,
        'veio_de_busca': bool(codigo_preenchido and endereco_retorno)
    }
    
    return render(request, 'produtos/cadastrar_produto.html', context)

@login_required
def editar_produto(request, produto_id):
    produto = get_object_or_404(Produto, id=produto_id)
    
    if request.method == 'POST':
        form = ProdutoForm(request.POST, instance=produto)
        if form.is_valid():
            form.save()
            messages.success(request, f'Produto "{produto.nome}" editado com sucesso!')
            return redirect('listar_produtos')
    else:
        form = ProdutoForm(instance=produto)
    
    return render(request, 'produtos/editar_produto.html', {
        'form': form, 
        'produto': produto
    })

@login_required
def listar_produtos(request):
    """Lista todos os produtos com opções de busca e edição"""
    search = request.GET.get('search', '')
    codigo = request.GET.get('codigo', '')
    status = request.GET.get('status', '')
    
    produtos = Produto.objects.all().order_by('nome')
    
    # Filtros de busca
    if search:
        produtos = produtos.filter(
            Q(nome__icontains=search) | 
            Q(categoria__icontains=search) |
            Q(fornecedor__icontains=search)
        )
    
    if codigo:
        produtos = produtos.filter(codigo__icontains=codigo)
    
    # Adicionar informações extras para cada produto
    produtos_processados = []
    for produto in produtos:
        # Contar endereços onde está armazenado
        total_enderecos = Estoque.objects.filter(produto=produto).count()
        
        # Contar lotes
        total_lotes = produto.lotes.count()
        
        # Próxima validade
        lotes_ordenados = produto.lotes.order_by('validade')
        proxima_validade = None
        dias_para_vencer = None
        
        if lotes_ordenados.exists():
            primeiro_lote = lotes_ordenados.first()
            proxima_validade = primeiro_lote.validade
            
            from datetime import date
            hoje = date.today()
            dias_para_vencer = (proxima_validade - hoje).days
        
        # Adicionar atributos ao produto
        produto.total_enderecos = total_enderecos
        produto.total_lotes = total_lotes
        produto.proxima_validade = proxima_validade
        produto.dias_para_vencer = dias_para_vencer
        
        produtos_processados.append(produto)
    
    # Filtrar por status se especificado
    if status:
        produtos_filtrados = []
        for produto in produtos_processados:
            if status == 'com_endereco' and produto.total_enderecos > 0:
                produtos_filtrados.append(produto)
            elif status == 'sem_endereco' and produto.total_enderecos == 0:
                produtos_filtrados.append(produto)
            elif status == 'com_lotes' and produto.total_lotes > 0:
                produtos_filtrados.append(produto)
            elif status == 'sem_lotes' and produto.total_lotes == 0:
                produtos_filtrados.append(produto)
        produtos_processados = produtos_filtrados
    
    # Calcular estatísticas
    total_produtos = len(produtos_processados)
    produtos_com_estoque = sum(1 for p in produtos_processados if p.total_enderecos > 0)
    produtos_sem_estoque = total_produtos - produtos_com_estoque
    produtos_com_lotes = sum(1 for p in produtos_processados if p.total_lotes > 0)
    
    context = {
        'produtos': produtos_processados,
        'produtos_com_estoque': produtos_com_estoque,
        'produtos_sem_estoque': produtos_sem_estoque,
        'produtos_com_lotes': produtos_com_lotes,
        'total_produtos': total_produtos
    }
    
    return render(request, 'produtos/listar_produtos.html', context)

@login_required
def excluir_produto(request, produto_id):
    produto = get_object_or_404(Produto, id=produto_id)
    
    # Verificar dados relacionados
    tem_estoque = Estoque.objects.filter(produto=produto).exists()
    tem_lotes = Lote.objects.filter(produto=produto).exists()
    tem_historico = HistoricoMovimentacao.objects.filter(produto=produto).exists()
    
    if request.method == 'GET':
        # Exibir confirmação via modal (já implementado no template)
        # Redirecionar para exclusão via GET
        try:
            # Excluir todos os dados relacionados
            count_estoque = Estoque.objects.filter(produto=produto).count()
            count_lotes = Lote.objects.filter(produto=produto).count()
            count_historico = HistoricoMovimentacao.objects.filter(produto=produto).count()
            
            # Deletar em cascata
            Estoque.objects.filter(produto=produto).delete()
            Lote.objects.filter(produto=produto).delete()
            HistoricoMovimentacao.objects.filter(produto=produto).delete()
            
            nome_produto = produto.nome
            codigo_produto = produto.codigo
            produto.delete()
            
            # Mensagem detalhada
            detalhes = []
            if count_estoque > 0:
                detalhes.append(f"{count_estoque} registro(s) de estoque")
            if count_lotes > 0:
                detalhes.append(f"{count_lotes} lote(s)")
            if count_historico > 0:
                detalhes.append(f"{count_historico} registro(s) de histórico")
            
            if detalhes:
                messages.success(request, f'Produto "{nome_produto}" (código: {codigo_produto}) e todos os dados relacionados foram excluídos com sucesso! Removido: {", ".join(detalhes)}.')
            else:
                messages.success(request, f'Produto "{nome_produto}" (código: {codigo_produto}) excluído com sucesso!')
            
        except Exception as e:
            messages.error(request, f'Erro ao excluir produto: {str(e)}')
    
    return redirect('listar_produtos')

@login_required
def cadastrar_enderecos(request):
    ultimo_endereco = None
    
    if request.method == 'POST':
        modo = request.POST.get('modo', '')
        
        # Modo de exclusão individual
        if modo == 'excluir':
            endereco_id = request.POST.get('endereco_id')
            if endereco_id:
                try:
                    endereco = Armazenamento.objects.get(id=endereco_id)
                    endereco.delete()
                    messages.success(request, f'Endereço {endereco.codigo} excluído com sucesso!')
                except Armazenamento.DoesNotExist:
                    messages.error(request, 'Endereço não encontrado!')
        
        # Modo de exclusão em lote
        elif modo == 'bulk_delete':
            endereco_ids = request.POST.getlist('endereco_ids')
            if endereco_ids:
                count = Armazenamento.objects.filter(id__in=endereco_ids).count()
                Armazenamento.objects.filter(id__in=endereco_ids).delete()
                messages.success(request, f'{count} endereço(s) excluído(s) com sucesso!')
        
        # Modo de alteração de categoria em lote
        elif modo == 'bulk_category':
            endereco_ids = request.POST.getlist('endereco_ids')
            nova_categoria = request.POST.get('nova_categoria')
            if endereco_ids and nova_categoria:
                count = Armazenamento.objects.filter(id__in=endereco_ids).update(categoria_armazenamento=nova_categoria)
                messages.success(request, f'Categoria alterada para {count} endereço(s)!')
        
        # Modo de geração em lote
        elif modo == 'lote':
            try:
                rua = int(request.POST.get('rua', 1))
                predio = int(request.POST.get('predio', 1))
                niveis = int(request.POST.get('niveis', 3))
                aps_por_nivel = int(request.POST.get('aps_por_nivel', 10))
                categoria = request.POST.get('categoria_lote', 'meio')
                
                enderecos_criados = []
                for nivel in range(niveis):
                    for ap in range(1, aps_por_nivel + 1):
                        codigo = f"R{rua:02d}P{predio:02d}N{nivel:02d}AP{ap:02d}"
                        
                        # Verificar se já existe
                        if not Armazenamento.objects.filter(codigo=codigo).exists():
                            endereco = Armazenamento.objects.create(
                                rua=str(rua),
                                predio=str(predio),
                                nivel=nivel,
                                ap=ap,
                                categoria_armazenamento=categoria,
                                codigo=codigo
                            )
                            enderecos_criados.append(endereco)
                
                messages.success(request, f'{len(enderecos_criados)} endereços criados com sucesso!')
                
            except (ValueError, TypeError) as e:
                messages.error(request, f'Erro nos dados fornecidos: {str(e)}')
        
        # Modo individual
        elif modo == 'individual':
            try:
                rua = request.POST.get('rua_individual')
                predio = request.POST.get('predio_individual')
                nivel = int(request.POST.get('nivel_individual'))
                ap = int(request.POST.get('ap_individual'))
                categoria = request.POST.get('categoria_individual')
                
                codigo = f"R{int(rua):02d}P{int(predio):02d}N{nivel:02d}AP{ap:02d}"
                
                # Verificar se já existe
                if Armazenamento.objects.filter(codigo=codigo).exists():
                    messages.error(request, f'Endereço {codigo} já existe!')
                else:
                    endereco = Armazenamento.objects.create(
                        rua=rua,
                        predio=predio,
                        nivel=nivel,
                        ap=ap,
                        categoria_armazenamento=categoria,
                        codigo=codigo
                    )
                    messages.success(request, f'Endereço {codigo} cadastrado com sucesso!')
                    ultimo_endereco = endereco
                    
            except (ValueError, TypeError) as e:
                messages.error(request, f'Erro nos dados fornecidos: {str(e)}')
        
        # Fallback para formulário antigo
        else:
            form = ArmazenamentoForm(request.POST)
            if form.is_valid():
                ultimo_endereco = form.save()
                messages.success(request, 'Endereço cadastrado com sucesso!')
    
    # Buscar todos os endereços e calcular estatísticas
    todos_enderecos = Armazenamento.objects.all().order_by('rua', 'predio', 'nivel', 'ap')
    
    # Calcular estatísticas
    total_enderecos = todos_enderecos.count()
    enderecos_com_estoque = Estoque.objects.values('local').distinct().count()
    enderecos_ocupados = enderecos_com_estoque
    enderecos_vazios = total_enderecos - enderecos_ocupados
    taxa_ocupacao = round((enderecos_ocupados / total_enderecos) * 100, 1) if total_enderecos > 0 else 0
    
    # Adicionar informações extras aos endereços
    enderecos_processados = []
    for endereco in todos_enderecos:
        total_produtos = Estoque.objects.filter(local=endereco).count()
        endereco.total_produtos = total_produtos
        endereco.tem_produtos = total_produtos > 0
        enderecos_processados.append(endereco)
    
    # Formulário para casos antigos
    form = ArmazenamentoForm()
    
    context = {
        'form': form,
        'ultimo_endereco': ultimo_endereco,
        'enderecos': enderecos_processados,
        'total_enderecos': total_enderecos,
        'enderecos_ocupados': enderecos_ocupados,
        'enderecos_vazios': enderecos_vazios,
        'taxa_ocupacao': taxa_ocupacao,
    }
    
    return render(request, 'produtos/cadastrar_enderecos.html', context)

@login_required
def exportar_estoque_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="estoque.csv"'
    writer = csv.writer(response)
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from django.utils.encoding import smart_str
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="estoque.xlsx"'
    wb = Workbook()
    ws = wb.active
    ws.title = "Estoque"
    # Cabeçalho
    headers = ['Rua', 'Prédio', 'Nível', 'Ap', 'Código', 'Produto', 'Validade', 'Data de Armazenamento']
    ws.append(headers)
    # Dados
    all_rows = [headers]
    for item in Estoque.objects.select_related('produto', 'local'):
        lotes = list(item.produto.lotes.all().order_by('validade'))
        if lotes:
            for lote in lotes:
                row = [
                    str(item.local.rua),
                    str(item.local.predio),
                    str(item.local.nivel),
                    str(item.local.ap),
                    str(item.produto.codigo),
                    str(item.produto.nome),
                    lote.validade.strftime('%d/%m/%Y'),
                    item.data_armazenado.strftime('%d/%m/%Y')
                ]
                ws.append(row)
                all_rows.append(row)
        else:
            row = [
                str(item.local.rua),
                str(item.local.predio),
                str(item.local.nivel),
                str(item.local.ap),
                str(item.produto.codigo),
                str(item.produto.nome),
                '-',
                item.data_armazenado.strftime('%d/%m/%Y')
            ]
            ws.append(row)
            all_rows.append(row)
    # Ajusta largura das colunas conforme maior conteúdo
    for col_idx, col in enumerate(zip(*all_rows), 1):
        max_length = max(len(str(cell)) for cell in col)
        ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    response.write(output.getvalue())
    return response

@login_required
def importar_produtos_csv(request):
    """View para importar produtos via CSV"""
    if request.method == 'POST':
        if 'csv_file' not in request.FILES:
            messages.error(request, 'Nenhum arquivo foi enviado.')
            return redirect('importar_produtos_csv')
        
        csv_file = request.FILES['csv_file']
        
        if not csv_file.name.endswith('.csv'):
            messages.error(request, 'Por favor, envie um arquivo CSV válido.')
            return redirect('importar_produtos_csv')
        
        try:
            # Lê o arquivo CSV
            decoded_file = csv_file.read().decode('utf-8-sig')  # utf-8-sig para remover BOM se presente
            io_string = io.StringIO(decoded_file)
            reader = csv.DictReader(io_string)
            
            produtos_criados = 0
            produtos_existentes = 0
            erros = []
            
            # Mapear possíveis nomes de colunas
            mapeamento_colunas = {
                # Possíveis nomes para código
                'codigo': ['codigo', 'código', 'code', 'id', 'item', 'cod', 'cod_produto', 'produto_codigo'],
                # Possíveis nomes para nome
                'nome': ['nome', 'name', 'produto', 'descricao', 'descrição', 'description', 'item_name', 'produto_nome']
            }
            
            # Detectar as colunas corretas no CSV
            colunas_detectadas = list(reader.fieldnames)
            coluna_codigo = None
            coluna_nome = None
            
            for coluna in colunas_detectadas:
                coluna_lower = coluna.lower().strip()
                
                # Detectar coluna de código
                if not coluna_codigo:
                    for possivel_nome in mapeamento_colunas['codigo']:
                        if possivel_nome in coluna_lower:
                            coluna_codigo = coluna
                            break
                
                # Detectar coluna de nome
                if not coluna_nome:
                    for possivel_nome in mapeamento_colunas['nome']:
                        if possivel_nome in coluna_lower:
                            coluna_nome = coluna
                            break
            
            if not coluna_codigo or not coluna_nome:
                messages.error(request, f'Não foi possível detectar as colunas de código e nome. Colunas encontradas: {", ".join(colunas_detectadas)}')
                return redirect('importar_produtos_csv')
            
            # Processar cada linha
            for linha_num, row in enumerate(reader, start=2):  # Começar em 2 porque linha 1 é o cabeçalho
                try:
                    codigo = str(row[coluna_codigo]).strip()
                    nome = str(row[coluna_nome]).strip()
                    
                    if not codigo or not nome or codigo.lower() in ['nan', 'null', ''] or nome.lower() in ['nan', 'null', '']:
                        continue
                    
                    # Verificar se o produto já existe
                    if Produto.objects.filter(codigo=codigo).exists():
                        produtos_existentes += 1
                        continue
                    
                    # Criar o produto
                    Produto.objects.create(
                        codigo=codigo,
                        nome=nome,
                        peso='',  # Campo obrigatório, mas pode ficar vazio
                        categoria='Importado CSV',
                        fornecedor='',
                    )
                    
                    produtos_criados += 1
                    
                except Exception as e:
                    erros.append(f'Linha {linha_num}: {str(e)}')
            
            # Mensagens de resultado
            if produtos_criados > 0:
                messages.success(request, f'✅ {produtos_criados} produtos foram importados com sucesso!')
            
            if produtos_existentes > 0:
                messages.info(request, f'ℹ️ {produtos_existentes} produtos já existiam no sistema.')
            
            if erros:
                messages.warning(request, f'⚠️ {len(erros)} erros encontrados durante a importação.')
                for erro in erros[:5]:  # Mostrar apenas os primeiros 5 erros
                    messages.error(request, erro)
                if len(erros) > 5:
                    messages.error(request, f'... e mais {len(erros) - 5} erros.')
            
            return redirect('importar_produtos_csv')
            
        except Exception as e:
            messages.error(request, f'Erro ao processar o arquivo: {str(e)}')
            return redirect('importar_produtos_csv')
    
    # GET request - mostrar formulário
    total_produtos = Produto.objects.count()
    return render(request, 'produtos/importar_csv.html', {
        'total_produtos': total_produtos
    })

@login_required
def importar_abastecimento_csv(request):
    """View para importar dados de abastecimento (estoque) via CSV"""
    if request.method == 'POST':
        if 'csv_file' not in request.FILES:
            messages.error(request, 'Nenhum arquivo foi enviado.')
            return redirect('importar_abastecimento_csv')
        
        csv_file = request.FILES['csv_file']
        
        if not csv_file.name.endswith('.csv'):
            messages.error(request, 'Por favor, envie um arquivo CSV válido.')
            return redirect('importar_abastecimento_csv')
        
        try:
            # Lê o arquivo CSV
            decoded_file = csv_file.read().decode('utf-8-sig')
            io_string = io.StringIO(decoded_file)
            
            # Detectar separador (vírgula ou ponto-e-vírgula)
            separador = ','
            if ';' in decoded_file and decoded_file.count(';') > decoded_file.count(','):
                separador = ';'
                
            # Reset do io_string com separador correto
            io_string = io.StringIO(decoded_file)
            reader = csv.DictReader(io_string, delimiter=separador)
            
            produtos_abastecidos = 0
            produtos_nao_encontrados = 0
            produtos_autocadastrados = 0
            produtos_sem_validade = 0
            enderecos_criados = 0
            erros = []
            
            # Mapear possíveis nomes de colunas - Melhorado para FIFO
            mapeamento_colunas = {
                'codigo': ['codigo', 'código', 'code', 'id', 'item', 'cod', 'cod_produto', 'produto_codigo', 'cód', 'cód.', 'item_code', 'sku'],
                'nome': ['nome', 'name', 'produto', 'descricao', 'descrição', 'description', 'item_name', 'produto_nome', 'item_descricao'],
                'rua': ['rua', 'street', 'endereco_rua', 'rua_armazenamento', 'corredor', 'setor', 'area', 'zona'],
                'predio': ['predio', 'prédio', 'building', 'pred', 'endereco_predio', 'edificio', 'bloco', 'galpao', 'galpão'],
                'nivel': ['nivel', 'nível', 'floor', 'andar', 'endereco_nivel', 'pavimento', 'piso', 'niv', 'niv.'],
                'ap': ['ap', 'apartamento', 'apt', 'endereco_ap', 'posicao', 'posição', 'endereco', 'endereço', 'local', 'sala', 'box'],
                'validade': ['validade', 'vencimento', 'expiration', 'data_validade', 'expire_date', 'data_vencimento', 'vence_em', 'exp_date', 'val', 'val.', 'val. 1', 'val. 2', 'val. 3', 'val. 4'],
                'data_armazenado': ['data_armazenado', 'data_armazenamento', 'data_entrada', 'storage_date', 'entrada', 'data_estoque', 'recebimento', 'data abastec', 'data abastec.'],
                'lote': ['lote', 'numero_lote', 'batch', 'lot_number', 'numero_batch', 'batch_number', 'lote_numero', 'num_lote'],
                'quantidade': ['quantidade', 'qty', 'qtd', 'quantity', 'amount', 'estoque', 'saldo', 'unidades', 'qte'],
                'tipo': ['tipo', 'type', 'categoria', 'class', 'classificacao'],
                'palete': ['palete', 'pallet', 'estrado', 'base']
            }
            
            # Detectar as colunas corretas no CSV - Melhorado para FIFO
            colunas_detectadas = list(reader.fieldnames)
            colunas_mapeadas = {}
            
            # Primeiro, limpar e normalizar os nomes das colunas
            colunas_normalizadas = {}
            for coluna in colunas_detectadas:
                # Remove caracteres especiais e normaliza (mantém espaços para casos como "val. 1")
                coluna_normalizada = coluna.lower().strip()
                # Remove pontos no final, mas mantém pontos internos para casos como "val. 1"
                if coluna_normalizada.endswith('.'):
                    coluna_normalizada = coluna_normalizada[:-1].strip()
                colunas_normalizadas[coluna_normalizada] = coluna
            
            # Mapear colunas com busca mais flexível
            for tipo_campo, possivel_nomes in mapeamento_colunas.items():
                for coluna_normalizada, coluna_original in colunas_normalizadas.items():
                    for possivel_nome in possivel_nomes:
                        possivel_nome_normalizado = possivel_nome.lower().strip()
                        # Busca exata primeiro, depois contida
                        if possivel_nome_normalizado == coluna_normalizada or possivel_nome_normalizado in coluna_normalizada or coluna_normalizada in possivel_nome_normalizado:
                            colunas_mapeadas[tipo_campo] = coluna_original
                            break
                    if tipo_campo in colunas_mapeadas:
                        break
            
            # Tratamento especial para múltiplas colunas de validade (VAL. 1, VAL. 2, etc.)
            if 'validade' not in colunas_mapeadas:
                for coluna in colunas_detectadas:
                    coluna_lower = coluna.lower().strip()
                    if 'val' in coluna_lower and any(c.isdigit() for c in coluna_lower):
                        colunas_mapeadas['validade'] = coluna
                        break
            
            # Se não encontrou todas as colunas obrigatórias, tentar estratégia alternativa
            colunas_obrigatorias = ['codigo', 'rua', 'predio', 'nivel', 'ap']
            faltando = [col for col in colunas_obrigatorias if col not in colunas_mapeadas]
            
            # Estratégia alternativa: Se temos apenas 2 colunas e uma parece ser código, assumir formato simples
            if len(colunas_detectadas) >= 2 and 'codigo' not in colunas_mapeadas:
                # Tentar detectar coluna de código pela posição ou padrão
                for i, coluna in enumerate(colunas_detectadas):
                    coluna_lower = coluna.lower().strip()
                    if any(palavra in coluna_lower for palavra in ['cod', 'cód', 'id', 'sku', 'item']):
                        colunas_mapeadas['codigo'] = coluna
                        break
                # Se ainda não encontrou, usar primeira coluna como código
                if 'codigo' not in colunas_mapeadas and colunas_detectadas:
                    colunas_mapeadas['codigo'] = colunas_detectadas[0]
            
            # Para CSV simples (só código/produto), definir endereços padrão
            usar_endereco_padrao = False
            if len(faltando) >= 3:  # Faltam pelo menos 3 campos de endereço
                usar_endereco_padrao = True
                # Valores padrão para endereçamento
                endereco_padrao = {
                    'rua': 'RUA FIFO',
                    'predio': 'PREDIO FIFO',  
                    'nivel': 'N1',
                    'ap': '01'
                }
                messages.info(request, f'ℹ️ Usando endereçamento automático FIFO: {endereco_padrao}')
            
            # Verificar se ainda faltam colunas essenciais
            if not usar_endereco_padrao:
                faltando = [col for col in colunas_obrigatorias if col not in colunas_mapeadas]
                if faltando:
                    messages.error(request, f'Colunas obrigatórias não encontradas: {", ".join(faltando)}. Colunas disponíveis: {", ".join(colunas_detectadas)}. Mapeamento detectado: {colunas_mapeadas}')
                    return redirect('importar_abastecimento_csv')
            
            # Processar cada linha - Melhorado para FIFO
            for linha_num, row in enumerate(reader, start=2):
                try:
                    codigo = str(row[colunas_mapeadas['codigo']]).strip()
                    
                    # Limpar código (remover caracteres especiais, espaços extras)
                    codigo = ''.join(c for c in codigo if c.isalnum() or c in '-_').strip()
                    
                    if not codigo or codigo.lower() in ['nan', 'null', '', 'none']:
                        continue
                    
                    # Buscar o produto ou criar se não existir
                    try:
                        produto = Produto.objects.get(codigo=codigo)
                    except Produto.DoesNotExist:
                        # Tentar extrair nome do produto se disponível
                        nome_produto = codigo  # Padrão: usar código como nome
                        if 'nome' in colunas_mapeadas and colunas_mapeadas['nome'] in row:
                            nome_str = str(row[colunas_mapeadas['nome']]).strip()
                            if nome_str and nome_str.lower() not in ['nan', 'null', '', 'none', '-']:
                                nome_produto = nome_str[:100]  # Limitar tamanho do nome
                        
                        # Criar produto automaticamente
                        try:
                            produto = Produto.objects.create(
                                codigo=codigo,
                                nome=nome_produto,
                                categoria='Importado FIFO',
                                peso='',
                                fornecedor='Auto-cadastrado'
                            )
                            produtos_autocadastrados += 1
                            messages.info(request, f'✨ Produto "{codigo} - {nome_produto}" foi cadastrado automaticamente')
                        except Exception as e:
                            produtos_nao_encontrados += 1
                            erros.append(f'Linha {linha_num}: Erro ao criar produto "{codigo}": {str(e)}')
                            continue
                    
                    # Extrair dados do endereço (usar padrão se necessário)
                    if usar_endereco_padrao:
                        rua = endereco_padrao['rua']
                        predio = endereco_padrao['predio']
                        nivel = endereco_padrao['nivel']
                        # Gerar AP único baseado no código do produto
                        ap = f"{endereco_padrao['ap']}-{codigo[:8]}"  # Aumentar limite do AP
                    else:
                        rua = str(row[colunas_mapeadas['rua']]).strip()
                        predio = str(row[colunas_mapeadas['predio']]).strip()
                        nivel = str(row[colunas_mapeadas['nivel']]).strip()
                        ap = str(row[colunas_mapeadas['ap']]).strip()
                    
                    # Validar dados do endereço
                    if not all([rua, predio, nivel, ap]):
                        if usar_endereco_padrao:
                            erros.append(f'Linha {linha_num}: Erro ao gerar endereço padrão para código "{codigo}"')
                        else:
                            erros.append(f'Linha {linha_num}: Dados de endereço incompletos para código "{codigo}"')
                        continue
                    
                    # Buscar ou criar endereço de armazenamento - Corrigido
                    try:
                        armazenamento = Armazenamento.objects.filter(
                            rua=rua,
                            predio=predio,
                            nivel=nivel,
                            ap=ap
                        ).first()
                        
                        if armazenamento:
                            created = False
                        else:
                            armazenamento = Armazenamento.objects.create(
                                rua=rua,
                                predio=predio,
                                nivel=nivel,
                                ap=ap,
                                livre=False,
                                capacidade_maxima=1
                            )
                            created = True
                    except Exception as e:
                        erros.append(f'Linha {linha_num}: Erro ao criar endereço: {str(e)}')
                        continue
                    
                    if created:
                        enderecos_criados += 1
                    else:
                        armazenamento.livre = False
                        armazenamento.save()
                    
                    # Extrair dados opcionais - Melhorado para diferentes formatos e múltiplas validades
                    validade = None
                    
                    # Tentar múltiplas colunas de validade (VAL. 1, VAL. 2, VAL. 3, VAL. 4)
                    colunas_validade_possiveis = []
                    if 'validade' in colunas_mapeadas:
                        colunas_validade_possiveis.append(colunas_mapeadas['validade'])
                    
                    # Adicionar outras colunas que contenham "val"
                    for coluna in colunas_detectadas:
                        coluna_lower = coluna.lower().strip()
                        if 'val' in coluna_lower and coluna not in colunas_validade_possiveis:
                            colunas_validade_possiveis.append(coluna)
                    
                    # Tentar extrair validade de qualquer uma das colunas
                    valores_validade_tentativas = {}
                    for coluna_val in colunas_validade_possiveis:
                        if coluna_val in row and row[coluna_val]:
                            try:
                                validade_str = str(row[coluna_val]).strip()
                                valores_validade_tentativas[coluna_val] = validade_str
                                
                                if validade_str and validade_str.lower() not in ['nan', 'null', '', 'none', '-']:
                                    # Tentar diferentes formatos de data
                                    formatos_data = [
                                        '%Y-%m-%d',     # 2025-12-31
                                        '%d/%m/%Y',     # 31/12/2025
                                        '%d-%m-%Y',     # 31-12-2025
                                        '%Y/%m/%d',     # 2025/12/31
                                        '%d.%m.%Y',     # 31.12.2025
                                        '%Y.%m.%d',     # 2025.12.31
                                        '%d/%m/%y',     # 31/12/25
                                        '%d-%m-%y',     # 31-12-25
                                        '%y/%m/%d',     # 25/12/31
                                        '%Y%m%d',       # 20251231
                                        '%d%m%Y',       # 31122025
                                        '%d%m%y',       # 311225
                                    ]
                                    
                                    for formato in formatos_data:
                                        try:
                                            validade = timezone.datetime.strptime(validade_str, formato).date()
                                            break
                                        except ValueError:
                                            continue
                                            
                                    if validade:
                                        break  # Encontrou validade válida, parar de procurar
                                        
                            except Exception as e:
                                continue  # Tentar próxima coluna
                    
                    # Só gerar erro se encontrou valores nas colunas mas nenhum era válido
                    if not validade and valores_validade_tentativas:
                        # Verificar se pelo menos uma coluna tinha um valor não vazio
                        tem_valor_nao_vazio = any(
                            valor and str(valor).strip() and str(valor).strip().lower() not in ['nan', 'null', '', 'none', '-']
                            for valor in valores_validade_tentativas.values()
                        )
                        
                        if tem_valor_nao_vazio:
                            # Limitar a 3 valores para não poluir o log
                            valores_limitados = dict(list(valores_validade_tentativas.items())[:3])
                            print(f'Aviso - Linha {linha_num}: Formato de data inválido nas colunas de validade. Valores: {valores_limitados}')
                    
                    data_armazenado = date.today()
                    if 'data_armazenado' in colunas_mapeadas and colunas_mapeadas['data_armazenado'] in row and row[colunas_mapeadas['data_armazenado']]:
                        try:
                            data_str = str(row[colunas_mapeadas['data_armazenado']]).strip()
                            if data_str and data_str.lower() not in ['nan', 'null', '', 'none', '-']:
                                for formato in formatos_data:
                                    try:
                                        data_armazenado = timezone.datetime.strptime(data_str, formato).date()
                                        break
                                    except ValueError:
                                        continue
                        except Exception as e:
                            erros.append(f'Linha {linha_num}: Erro ao processar data de armazenamento: {str(e)}')
                    
                    numero_lote = ''
                    if 'lote' in colunas_mapeadas and colunas_mapeadas['lote'] in row:
                        numero_lote = str(row[colunas_mapeadas['lote']]).strip()
                        if numero_lote.lower() in ['nan', 'null', '', 'none', '-']:
                            numero_lote = ''
                    
                    # Se não tem lote definido, gerar um baseado na data e código
                    if not numero_lote and validade:
                        numero_lote = f"FIFO-{codigo}-{validade.strftime('%Y%m%d')}"
                    elif not numero_lote:
                        numero_lote = f"FIFO-{codigo}-{data_armazenado.strftime('%Y%m%d')}"
                    
                    quantidade = 1
                    if 'quantidade' in colunas_mapeadas and colunas_mapeadas['quantidade'] in row:
                        try:
                            qty_str = str(row[colunas_mapeadas['quantidade']]).strip()
                            if qty_str and qty_str.lower() not in ['nan', 'null', '', 'none', '-']:
                                # Limpar e converter quantidade
                                qty_str = qty_str.replace(',', '.')  # Converter vírgula para ponto
                                quantidade = max(1, int(float(qty_str)))  # Garantir pelo menos 1
                        except (ValueError, TypeError):
                            quantidade = 1
                    
                    # Criar registro de estoque - Corrigido para evitar duplicatas
                    try:
                        # Tentar encontrar um registro existente
                        estoque = Estoque.objects.filter(
                            produto=produto,
                            local=armazenamento
                        ).first()
                        
                        if estoque:
                            # Se encontrou, usar o existente
                            estoque_created = False
                        else:
                            # Se não encontrou, criar novo
                            estoque = Estoque.objects.create(
                                produto=produto,
                                local=armazenamento,
                                data_armazenado=data_armazenado,
                                usuario_responsavel=request.user.username if request.user.is_authenticated else 'Sistema',
                                observacoes=f'Importado via CSV - Linha {linha_num}'
                            )
                            estoque_created = True
                    except Exception as e:
                        erros.append(f'Linha {linha_num}: Erro ao criar estoque: {str(e)}')
                        continue
                    
                    # Se não foi criado, atualizar data
                    if not estoque_created:
                        estoque.data_armazenado = data_armazenado
                        estoque.save()
                    
                    # Criar lote se tiver validade
                    if validade:
                        lote, lote_created = Lote.objects.get_or_create(
                            produto=produto,
                            validade=validade,
                            numero_lote=numero_lote,
                            defaults={'quantidade': quantidade}
                        )
                        
                        # Se lote já existe, somar quantidade
                        if not lote_created:
                            lote.quantidade += quantidade
                            lote.save()
                        
                        # Criar alertas se necessário
                        dias_para_vencer = (validade - date.today()).days
                        if dias_para_vencer <= 0:
                            Alerta.objects.get_or_create(
                                tipo='vencido',
                                produto=produto,
                                defaults={
                                    'titulo': f'Produto Vencido: {produto.nome}',
                                    'mensagem': f'O produto {produto.nome} (código {produto.codigo}) está vencido desde {validade.strftime("%d/%m/%Y")}'
                                }
                            )
                        elif dias_para_vencer <= 30:
                            Alerta.objects.get_or_create(
                                tipo='vencimento',
                                produto=produto,
                                defaults={
                                    'titulo': f'Produto Próximo ao Vencimento: {produto.nome}',
                                    'mensagem': f'O produto {produto.nome} (código {produto.codigo}) vence em {dias_para_vencer} dias ({validade.strftime("%d/%m/%Y")})'
                                }
                            )
                    else:
                        # Contar produtos sem validade (não é erro, é normal)
                        produtos_sem_validade += 1
                    
                    # Criar histórico de movimentação
                    HistoricoMovimentacao.objects.create(
                        produto=produto,
                        local_destino=armazenamento,
                        tipo_operacao='entrada',
                        quantidade=quantidade,
                        usuario=request.user.username if request.user.is_authenticated else 'Sistema CSV',
                        observacoes=f'Abastecimento via CSV - Lote: {numero_lote or "N/A"}'
                    )
                    
                    produtos_abastecidos += 1
                    
                except Exception as e:
                    erros.append(f'Linha {linha_num}: {str(e)}')
            
            # Mensagens de resultado
            if produtos_abastecidos > 0:
                messages.success(request, f'✅ {produtos_abastecidos} produtos foram abastecidos com sucesso!')
            
            if produtos_autocadastrados > 0:
                messages.success(request, f'✨ {produtos_autocadastrados} produtos foram cadastrados automaticamente!')
            
            if produtos_sem_validade > 0:
                messages.info(request, f'ℹ️ {produtos_sem_validade} produtos foram importados sem validade (normal para alguns produtos).')
            
            if enderecos_criados > 0:
                messages.info(request, f'🏢 {enderecos_criados} novos endereços de armazenamento foram criados.')
            
            if produtos_nao_encontrados > 0:
                messages.warning(request, f'⚠️ {produtos_nao_encontrados} produtos não puderam ser processados.')
            
            if erros and len(erros) <= 10:  # Mostrar todos os erros se forem poucos
                messages.warning(request, f'⚠️ {len(erros)} erros encontrados durante a importação.')
                for erro in erros:
                    messages.error(request, erro)
            elif erros:  # Muitos erros - mostrar apenas alguns
                messages.warning(request, f'⚠️ {len(erros)} erros encontrados durante a importação.')
                for erro in erros[:3]:  # Reduzir para 3 erros para não poluir
                    messages.error(request, erro)
                if len(erros) > 3:
                    messages.error(request, f'... e mais {len(erros) - 3} erros. Verifique os dados do CSV.')
            
            return redirect('importar_abastecimento_csv')
            
        except Exception as e:
            messages.error(request, f'Erro ao processar o arquivo: {str(e)}')
            return redirect('importar_abastecimento_csv')
    
    # GET request - mostrar formulário
    total_produtos = Produto.objects.count()
    total_estoque = Estoque.objects.count()
    total_enderecos = Armazenamento.objects.count()
    
    return render(request, 'produtos/importar_abastecimento.html', {
        'total_produtos': total_produtos,
        'total_estoque': total_estoque,
        'total_enderecos': total_enderecos
    })

@login_required
def alertas(request):
    """View para gerenciar alertas"""
    status_filtro = request.GET.get('status', 'ativo')
    
    alertas_list = Alerta.objects.all()
    if status_filtro:
        alertas_list = alertas_list.filter(status=status_filtro)
    
    alertas_list = alertas_list.order_by('-data_criacao')
    
    paginator = Paginator(alertas_list, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'produtos/alertas.html', {
        'page_obj': page_obj,
        'status_filtro': status_filtro
    })

@login_required
def marcar_alerta_lido(request, alerta_id):
    """Marca um alerta como lido"""
    alerta = get_object_or_404(Alerta, id=alerta_id)
    alerta.status = 'lido'
    alerta.data_leitura = timezone.now()
    alerta.save()
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': True})
    
    return redirect('alertas')

@login_required
def historico_movimentacoes(request):
    """View para histórico de movimentações"""
    tipo_filtro = request.GET.get('tipo', '')
    produto_filtro = request.GET.get('produto', '')
    
    movimentacoes = HistoricoMovimentacao.objects.select_related('produto', 'local_origem', 'local_destino')
    
    if tipo_filtro:
        movimentacoes = movimentacoes.filter(tipo_operacao=tipo_filtro)
    
    if produto_filtro:
        movimentacoes = movimentacoes.filter(
            Q(produto__nome__icontains=produto_filtro) |
            Q(produto__codigo__icontains=produto_filtro)
        )
    
    movimentacoes = movimentacoes.order_by('-data_operacao')
    
    paginator = Paginator(movimentacoes, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'produtos/historico.html', {
        'page_obj': page_obj,
        'tipo_filtro': tipo_filtro,
        'produto_filtro': produto_filtro,
        'tipos_operacao': HistoricoMovimentacao.TIPOS_OPERACAO
    })

@login_required
def extrair_produtos_abastecimento(request):
    """View para extrair e cadastrar produtos de um CSV de abastecimento"""
    if request.method == 'POST':
        if 'csv_file' not in request.FILES:
            messages.error(request, 'Nenhum arquivo foi enviado.')
            return redirect('extrair_produtos_abastecimento')
        
        csv_file = request.FILES['csv_file']
        
        if not csv_file.name.endswith('.csv'):
            messages.error(request, 'Por favor, envie um arquivo CSV válido.')
            return redirect('extrair_produtos_abastecimento')
        
        try:
            # Lê o arquivo CSV
            decoded_file = csv_file.read().decode('utf-8-sig')
            io_string = io.StringIO(decoded_file)
            
            # Detectar separador (vírgula ou ponto-e-vírgula)
            separador = ','
            if ';' in decoded_file and decoded_file.count(';') > decoded_file.count(','):
                separador = ';'
                
            # Reset do io_string com separador correto
            io_string = io.StringIO(decoded_file)
            reader = csv.DictReader(io_string, delimiter=separador)
            
            produtos_criados = 0
            produtos_existentes = 0
            erros = []
            
            # Mapear possíveis nomes de colunas
            mapeamento_colunas = {
                'codigo': ['codigo', 'código', 'code', 'id', 'item', 'cod', 'cod_produto', 'produto_codigo', 'cód', 'cód.', 'item_code', 'sku'],
                'nome': ['nome', 'name', 'produto', 'descricao', 'descrição', 'description', 'item_name', 'produto_nome', 'item_descricao'],
                'tipo': ['tipo', 'type', 'categoria', 'class', 'classificacao'],
            }
            
            # Detectar as colunas corretas no CSV
            colunas_detectadas = list(reader.fieldnames)
            colunas_mapeadas = {}
            
            # Normalizar e mapear colunas
            for coluna in colunas_detectadas:
                coluna_normalizada = coluna.lower().strip()
                if coluna_normalizada.endswith('.'):
                    coluna_normalizada = coluna_normalizada[:-1].strip()
                
                for tipo_campo, possivel_nomes in mapeamento_colunas.items():
                    for possivel_nome in possivel_nomes:
                        possivel_nome_normalizado = possivel_nome.lower().strip()
                        if possivel_nome_normalizado == coluna_normalizada or possivel_nome_normalizado in coluna_normalizada:
                            colunas_mapeadas[tipo_campo] = coluna
                            break
                    if tipo_campo in colunas_mapeadas:
                        break
            
            # Verificar se encontrou pelo menos código
            if 'codigo' not in colunas_mapeadas:
                # Tentar usar primeira coluna como código
                if colunas_detectadas:
                    colunas_mapeadas['codigo'] = colunas_detectadas[0]
                    messages.info(request, f'ℹ️ Usando "{colunas_detectadas[0]}" como coluna de código')
                else:
                    messages.error(request, 'Não foi possível detectar coluna de código no arquivo')
                    return redirect('extrair_produtos_abastecimento')
            
            # Processar cada linha
            for linha_num, row in enumerate(reader, start=2):
                try:
                    codigo = str(row[colunas_mapeadas['codigo']]).strip()
                    
                    # Limpar código
                    codigo = ''.join(c for c in codigo if c.isalnum() or c in '-_').strip()
                    
                    if not codigo or codigo.lower() in ['nan', 'null', '', 'none']:
                        continue
                    
                    # Verificar se o produto já existe
                    if Produto.objects.filter(codigo=codigo).exists():
                        produtos_existentes += 1
                        continue
                    
                    # Extrair nome do produto
                    nome_produto = codigo  # Padrão: usar código como nome
                    if 'nome' in colunas_mapeadas and colunas_mapeadas['nome'] in row:
                        nome_str = str(row[colunas_mapeadas['nome']]).strip()
                        if nome_str and nome_str.lower() not in ['nan', 'null', '', 'none', '-']:
                            nome_produto = nome_str[:100]  # Limitar tamanho do nome
                    
                    # Extrair categoria/tipo
                    categoria = 'Extraído de Abastecimento'
                    if 'tipo' in colunas_mapeadas and colunas_mapeadas['tipo'] in row:
                        tipo_str = str(row[colunas_mapeadas['tipo']]).strip()
                        if tipo_str and tipo_str.lower() not in ['nan', 'null', '', 'none', '-']:
                            categoria = tipo_str[:50]
                    
                    # Criar o produto
                    Produto.objects.create(
                        codigo=codigo,
                        nome=nome_produto,
                        categoria=categoria,
                        peso='',
                        fornecedor='Extraído CSV Abastecimento',
                    )
                    
                    produtos_criados += 1
                    
                except Exception as e:
                    erros.append(f'Linha {linha_num}: {str(e)}')
            
            # Mensagens de resultado
            if produtos_criados > 0:
                messages.success(request, f'✅ {produtos_criados} produtos foram extraídos e cadastrados com sucesso!')
            
            if produtos_existentes > 0:
                messages.info(request, f'ℹ️ {produtos_existentes} produtos já existiam no sistema.')
            
            if erros:
                messages.warning(request, f'⚠️ {len(erros)} erros encontrados durante a extração.')
                for erro in erros[:5]:
                    messages.error(request, erro)
                if len(erros) > 5:
                    messages.error(request, f'... e mais {len(erros) - 5} erros.')
            
            return redirect('extrair_produtos_abastecimento')
            
        except Exception as e:
            messages.error(request, f'Erro ao processar o arquivo: {str(e)}')
            return redirect('extrair_produtos_abastecimento')
    
    # GET request - mostrar formulário
    total_produtos = Produto.objects.count()
    return render(request, 'produtos/extrair_produtos_abastecimento.html', {
        'total_produtos': total_produtos
    })

def marcar_saida(request, estoque_id):
    """Marca um produto como saída, transferindo de 'inteiro' para 'meio'"""
    try:
        estoque = get_object_or_404(Estoque, id=estoque_id)
        
        # Verifica se o produto está atualmente como 'inteiro'
        if estoque.local.categoria_armazenamento != 'inteiro':
            messages.warning(request, 'Este produto já está marcado como saída.')
            return redirect('painel')
        
        # Procura por um local de saída (categoria 'meio') disponível
        # Preferencialmente na mesma rua e prédio
        local_saida = Armazenamento.objects.filter(
            categoria_armazenamento='meio',
            rua=estoque.local.rua,
            predio=estoque.local.predio,
            livre=True
        ).first()
        
        # Se não encontrou na mesma rua/prédio, procura qualquer local de saída disponível
        if not local_saida:
            local_saida = Armazenamento.objects.filter(
                categoria_armazenamento='meio',
                livre=True
            ).first()
        
        # Se ainda não encontrou, cria um novo local de saída no nível 0
        if not local_saida:
            local_saida = Armazenamento.objects.create(
                categoria_armazenamento='meio',
                rua=estoque.local.rua,
                predio=estoque.local.predio,
                nivel='0',
                ap=f'SAIDA-{estoque.local.rua}-{estoque.local.predio}',
                livre=True,
                capacidade_maxima=100,  # Capacidade maior para área de saída
                observacoes='Local de saída criado automaticamente'
            )
        
        # Registra a movimentação no histórico
        HistoricoMovimentacao.objects.create(
            produto=estoque.produto,
            tipo_operacao='saida',
            local_origem=estoque.local,
            local_destino=local_saida,
            data_operacao=date.today(),
            usuario_responsavel=request.user.username if request.user.is_authenticated else 'Sistema',
            observacoes=f'Produto marcado como saída - transferido de {estoque.local} para {local_saida}'
        )
        
        # Atualiza o local do estoque
        estoque_antigo = estoque.local
        estoque.local = local_saida
        estoque.data_armazenado = date.today()
        estoque.observacoes = f'Transferido para saída em {date.today().strftime("%d/%m/%Y")}'
        estoque.save()
        
        # Marca local antigo como livre se não houver mais produtos
        if not Estoque.objects.filter(local=estoque_antigo).exists():
            estoque_antigo.livre = True
            estoque_antigo.save()
        
        # Marca novo local como ocupado
        local_saida.livre = False
        local_saida.save()
        
        messages.success(request, f'Produto "{estoque.produto.nome}" marcado como saída com sucesso!')
        
    except Exception as e:
        messages.error(request, f'Erro ao marcar produto como saída: {str(e)}')
    
    return redirect('painel')

def alterar_tipo_endereco(request, endereco_id):
    """Altera o tipo de armazenamento de um endereço"""
    try:
        endereco = get_object_or_404(Armazenamento, id=endereco_id)
        
        # Verifica se há produtos neste endereço
        produtos_no_endereco = Estoque.objects.filter(local=endereco).count()
        
        if produtos_no_endereco > 0:
            messages.warning(request, f'Não é possível alterar o tipo. Este endereço possui {produtos_no_endereco} produto(s) armazenado(s).')
            return redirect('painel')
        
        # Alterna o tipo
        if endereco.categoria_armazenamento == 'inteiro':
            endereco.categoria_armazenamento = 'meio'
            tipo_novo = 'Saída (Nível 0)'
        else:
            endereco.categoria_armazenamento = 'inteiro'
            tipo_novo = 'Palete Fechado (Nível 2)'
        
        endereco.save()
        
        messages.success(request, f'Tipo do endereço alterado para "{tipo_novo}" com sucesso!')
        
    except Exception as e:
        messages.error(request, f'Erro ao alterar tipo do endereço: {str(e)}')
    
    return redirect('painel')

def alterar_tipos_lote(request):
    """Altera o tipo de armazenamento de múltiplos endereços"""
    if request.method == 'POST':
        endereco_ids = request.POST.getlist('endereco_ids')
        novo_tipo = request.POST.get('novo_tipo')
        
        if not endereco_ids:
            messages.error(request, 'Nenhum endereço foi selecionado.')
            return redirect('cadastrar_enderecos')
            
        if novo_tipo not in ['inteiro', 'meio']:
            messages.error(request, 'Tipo de armazenamento inválido.')
            return redirect('cadastrar_enderecos')
        
        # Verificar quais endereços têm produtos
        enderecos_com_produtos = []
        enderecos_vazios = []
        
        for endereco_id in endereco_ids:
            try:
                endereco = Armazenamento.objects.get(id=endereco_id)
                ocupacao = Estoque.objects.filter(local=endereco).count()
                
                if ocupacao > 0:
                    enderecos_com_produtos.append(f"{endereco} ({ocupacao} produtos)")
                else:
                    enderecos_vazios.append(endereco)
            except Armazenamento.DoesNotExist:
                continue
        
        # Alertar sobre endereços com produtos
        if enderecos_com_produtos:
            messages.warning(request, 
                f'Os seguintes endereços não foram alterados por possuírem produtos: {", ".join(enderecos_com_produtos[:3])}' +
                (f' e mais {len(enderecos_com_produtos) - 3}' if len(enderecos_com_produtos) > 3 else ''))
        
        # Alterar apenas os endereços vazios
        if enderecos_vazios:
            tipo_nome = 'Palete Fechado (Nível 2)' if novo_tipo == 'inteiro' else 'Saída (Nível 0)'
            
            for endereco in enderecos_vazios:
                endereco.categoria_armazenamento = novo_tipo
                endereco.save()
            
            messages.success(request, 
                f'{len(enderecos_vazios)} endereço(s) alterado(s) para "{tipo_nome}" com sucesso!')
        
        if not enderecos_vazios and not enderecos_com_produtos:
            messages.info(request, 'Nenhum endereço válido foi encontrado para alteração.')
    
    return redirect('cadastrar_enderecos')

def buscar_produto_endereco(request, endereco_id):
    """
    Busca produto por código para armazenar em endereço específico
    """
    endereco = get_object_or_404(Armazenamento, id=endereco_id)
    produto_encontrado = None
    erro = None
    
    if request.method == 'POST':
        codigo = request.POST.get('codigo', '').strip()
        if codigo:
            try:
                produto_encontrado = Produto.objects.get(codigo=codigo)
                # Verificar se já existe estoque neste endereço
                if Estoque.objects.filter(produto=produto_encontrado, local=endereco).exists():
                    erro = "Este produto já está armazenado neste endereço!"
                    produto_encontrado = None
            except Produto.DoesNotExist:
                # Redirecionar para cadastro de produto com código pré-preenchido
                return redirect(f'/produtos/cadastrar_produto/?codigo={codigo}&endereco_retorno={endereco.id}')
    
    context = {
        'endereco': endereco,
        'produto_encontrado': produto_encontrado,
        'erro': erro,
        'codigo_pesquisado': request.POST.get('codigo', '') if request.method == 'POST' else ''
    }
    
    return render(request, 'produtos/buscar_produto_endereco.html', context)

def confirmar_armazenamento_endereco(request, endereco_id, produto_id):
    """
    Confirma armazenamento do produto no endereço específico
    """
    endereco = get_object_or_404(Armazenamento, id=endereco_id)
    produto = get_object_or_404(Produto, id=produto_id)
    
    # Verificar se já existe estoque neste endereço
    if Estoque.objects.filter(produto=produto, local=endereco).exists():
        messages.error(request, "Este produto já está armazenado neste endereço!")
        return redirect('painel')
    
    # Valores padrão vindos do formulário rápido (GET)
    data_validade_quick = request.GET.get('data_validade_quick')
    data_armazenado_quick = request.GET.get('data_armazenado_quick')
    
    if request.method == 'POST':
        # Buscar múltiplas validades e quantidades
        validades = request.POST.getlist('data_validade')
        quantidades = request.POST.getlist('quantidade')
        observacoes = request.POST.get('observacoes', '')
        data_armazenado = request.POST.get('data_armazenado')
        
        # Garantir que temos pelo menos uma validade
        if not validades or not validades[0]:
            messages.error(request, 'É necessário informar pelo menos uma data de validade!')
            return render(request, 'produtos/confirmar_armazenamento_endereco.html', {
                'endereco': endereco,
                'produto': produto,
                'data_validade_default': data_validade_quick,
                'data_armazenado_default': data_armazenado_quick,
                'next': request.POST.get('next')
            })
        
        # Usar data de armazenamento fornecida ou data atual
        data_armazenado_final = data_armazenado if data_armazenado else timezone.now().date()
        
        produtos_armazenados = 0
        lotes_criados = []
        
        # Processar cada validade
        for i, validade in enumerate(validades):
            if not validade:  # Pular validades vazias
                continue
                
            quantidade = int(quantidades[i]) if i < len(quantidades) and quantidades[i] else 1
            
            try:
                # Buscar lote existente ou criar novo
                lote = Lote.objects.filter(
                    produto=produto,
                    validade=validade
                ).first()
                
                if not lote:
                    # Gerar número de lote único
                    base_numero = f'LOTE-{produto.codigo}-{validade.replace("-", "")}'
                    contador = 1
                    numero_lote = base_numero
                    
                    while Lote.objects.filter(numero_lote=numero_lote).exists():
                        numero_lote = f'{base_numero}-{contador}'
                        contador += 1
                    
                    lote = Lote.objects.create(
                        produto=produto,
                        validade=validade,
                        numero_lote=numero_lote,
                        quantidade=quantidade
                    )
                    lotes_criados.append(lote.numero_lote)
                else:
                    # Atualizar quantidade do lote existente
                    lote.quantidade += quantidade
                    lote.save()
                
                # Criar entradas de estoque para cada unidade
                for _ in range(quantidade):
                    Estoque.objects.create(
                        produto=produto,
                        local=endereco,
                        data_armazenado=data_armazenado_final,
                        data_validade=validade,
                        observacoes=observacoes or f'Armazenado via modal - Lote: {lote.numero_lote} em {timezone.now().strftime("%d/%m/%Y %H:%M")}',
                    )
                    produtos_armazenados += 1
                    
            except Exception as e:
                messages.error(request, f'Erro ao processar validade {validade}: {str(e)}')
                continue
        
        if produtos_armazenados > 0:
            messages.success(request, f'{produtos_armazenados} unidade(s) do produto "{produto.nome}" armazenada(s) com sucesso em {endereco}!')
            if lotes_criados:
                messages.info(request, f'Lotes criados: {", ".join(lotes_criados)}')
        else:
            messages.error(request, 'Nenhum produto foi armazenado. Verifique os dados informados.')
        
        # Verificar se veio da busca avançada para retornar com filtros
        next_url = request.POST.get('next') or request.GET.get('next')
        if next_url and 'busca-endereco-avancada' in next_url:
            return redirect(next_url)
        
        return redirect('painel')
    
    # Processar armazenamento rápido via GET (do formulário da busca)
    elif data_validade_quick:
        # Buscar múltiplas validades e quantidades
        validades = request.GET.getlist('data_validade_quick')
        quantidades = request.GET.getlist('quantidade')
        
        # Garantir que temos pelo menos uma validade
        if not validades or not validades[0]:
            messages.error(request, 'É necessário informar pelo menos uma data de validade!')
            return redirect('buscar_produto_endereco', endereco_id=endereco.id)
        
        # Usar data de armazenamento fornecida ou data atual
        data_armazenado_final = data_armazenado_quick or timezone.now().date()
        
        produtos_armazenados = 0
        lotes_criados = []
        
        # Processar cada validade
        for i, validade in enumerate(validades):
            if not validade:  # Pular validades vazias
                continue
                
            quantidade = int(quantidades[i]) if i < len(quantidades) and quantidades[i] else 1
            
            try:
                # Buscar lote existente ou criar novo
                lote = Lote.objects.filter(
                    produto=produto,
                    validade=validade
                ).first()
                
                if not lote:
                    # Gerar número de lote único
                    base_numero = f'LOTE-{produto.codigo}-{validade.replace("-", "")}'
                    contador = 1
                    numero_lote = base_numero
                    
                    while Lote.objects.filter(numero_lote=numero_lote).exists():
                        numero_lote = f'{base_numero}-{contador}'
                        contador += 1
                    
                    lote = Lote.objects.create(
                        produto=produto,
                        validade=validade,
                        numero_lote=numero_lote,
                        quantidade=quantidade
                    )
                    lotes_criados.append(lote.numero_lote)
                else:
                    # Atualizar quantidade do lote existente
                    lote.quantidade += quantidade
                    lote.save()
                
                # Criar entradas de estoque para cada unidade
                for _ in range(quantidade):
                    Estoque.objects.create(
                        produto=produto,
                        local=endereco,
                        data_armazenado=data_armazenado_final,
                        data_validade=validade,
                        observacoes=f'Armazenado via busca rápida - Lote: {lote.numero_lote} em {timezone.now().strftime("%d/%m/%Y %H:%M")}',
                    )
                    produtos_armazenados += 1
                    
            except Exception as e:
                messages.error(request, f'Erro ao processar validade {validade}: {str(e)}')
                continue
        
        if produtos_armazenados > 0:
            messages.success(request, f'{produtos_armazenados} unidade(s) do produto "{produto.nome}" armazenada(s) com sucesso em {endereco}!')
            if lotes_criados:
                messages.info(request, f'Lotes criados: {", ".join(lotes_criados)}')
        else:
            messages.error(request, 'Nenhum produto foi armazenado. Verifique os dados informados.')
        
        # Verificar se veio da busca avançada para retornar com filtros
        next_url = request.GET.get('next')
        if next_url and 'busca-endereco-avancada' in next_url:
            return redirect(next_url)
        
        return redirect('painel')
    
    context = {
        'endereco': endereco,
        'produto': produto,
        'data_validade_default': data_validade_quick,
        'data_armazenado_default': data_armazenado_quick
    }
    
    return render(request, 'produtos/confirmar_armazenamento_endereco.html', context)

@login_required
def pagina_principal(request):
    """
    Página principal que combina dashboard com painel de estoque completo.
    Inclui botões úteis, busca rápida e visualização completa do estoque.
    """
    # Processar busca por código
    busca_codigo = request.GET.get('busca_codigo', '')
    resultado_busca = None
    
    if busca_codigo:
        try:
            produto = Produto.objects.get(codigo=busca_codigo)
            
            # Adicionar informações de validade
            lotes = produto.lotes.all().order_by('validade')
            if lotes.exists():
                primeiro_lote = lotes.first()
                from datetime import date
                hoje = date.today()
                dias_para_vencer = (primeiro_lote.validade - hoje).days
                
                if dias_para_vencer < 0:
                    status_validade = "Vencido"
                elif dias_para_vencer <= 7:
                    status_validade = "Vence em breve"
                elif dias_para_vencer <= 30:
                    status_validade = "Próximo ao vencimento"
                else:
                    status_validade = "Válido"
                
                produto.status_validade = status_validade
                produto.dias_para_vencer = dias_para_vencer
                produto.proxima_validade = primeiro_lote.validade
            
            resultado_busca = produto
        except Produto.DoesNotExist:
            resultado_busca = None
    
    # Buscar todos os endereços cadastrados (não apenas com estoque) ordenados
    enderecos = Armazenamento.objects.all().order_by('rua', 'predio', 'nivel', 'ap')
    
    # Organizar por ruas e prédios usando lógica otimizada
    organizacao = {}
    
    for endereco in enderecos:
        rua = int(endereco.rua) if endereco.rua.isdigit() else endereco.rua
        predio = int(endereco.predio) if endereco.predio.isdigit() else endereco.predio
        
        if rua not in organizacao:
            organizacao[rua] = {}
        if predio not in organizacao[rua]:
            organizacao[rua][predio] = {'enderecos_com_produtos': [], 'enderecos_vazios': []}
        
        # Buscar produtos neste endereço
        estoques = Estoque.objects.filter(local=endereco).select_related('produto')
        produtos_com_lotes = []
        
        if estoques.exists():
            # Endereço com produtos
            for estoque in estoques:
                produto = estoque.produto
                lotes = produto.lotes.all().order_by('validade')
                
                # Calcular próxima validade e status
                proxima_validade = None
                status_validade = "Sem lote"
                dias_para_vencer = None
                
                if lotes.exists():
                    primeiro_lote = lotes.first()
                    proxima_validade = primeiro_lote.validade
                    
                    from datetime import date
                    hoje = date.today()
                    dias_para_vencer = (proxima_validade - hoje).days
                    
                    if dias_para_vencer < 0:
                        status_validade = "Vencido"
                    elif dias_para_vencer <= 7:
                        status_validade = "Vence em breve"
                    elif dias_para_vencer <= 30:
                        status_validade = "Próximo ao vencimento"
                    else:
                        status_validade = "Válido"
                
                produto_info = {
                    'estoque': estoque,
                    'produto': produto,
                    'lotes': lotes,
                    'proxima_validade': proxima_validade,
                    'status_validade': status_validade,
                    'dias_para_vencer': dias_para_vencer,
                    'total_lotes': lotes.count()
                }
                produtos_com_lotes.append(produto_info)
            
            # Adicionar informações do endereço com produtos
            endereco_info = {
                'endereco': endereco,
                'produtos': produtos_com_lotes,
                'tem_produtos': True,
                'total_produtos': len(produtos_com_lotes)
            }
            
            organizacao[rua][predio]['enderecos_com_produtos'].append(endereco_info)
        else:
            # Endereço vazio
            organizacao[rua][predio]['enderecos_vazios'].append(endereco)
    
    # Ordenar as ruas e prédios numericamente
    def ordenacao_inteligente(valor):
        """Ordena numericamente se possível, senão alfabeticamente"""
        try:
            return (0, int(valor))  # Se for número, usa ordenação numérica
        except ValueError:
            return (1, str(valor).lower())  # Se não for número, usa ordenação alfabética
    
    # Criar organizacao ordenada
    organizacao_ordenada = {}
    ruas_ordenadas = sorted(organizacao.keys(), key=ordenacao_inteligente)
    
    for rua in ruas_ordenadas:
        organizacao_ordenada[rua] = {}
        predios_ordenados = sorted(organizacao[rua].keys(), key=ordenacao_inteligente)
        
        for predio in predios_ordenados:
            # Calcular total de produtos no prédio
            total_produtos_predio = sum(endereco_info['total_produtos'] 
                                      for endereco_info in organizacao[rua][predio]['enderecos_com_produtos'])
            
            # Calcular total de endereços vazios no prédio
            total_vazios_predio = len(organizacao[rua][predio]['enderecos_vazios'])
            
            organizacao_ordenada[rua][predio] = {
                'enderecos': organizacao[rua][predio]['enderecos_com_produtos'],
                'enderecos_vazios': organizacao[rua][predio]['enderecos_vazios'],
                'total_produtos': total_produtos_predio,
                'total_vazios': total_vazios_predio
            }
    
    # Calcular estatísticas
    total_enderecos = Armazenamento.objects.count()
    enderecos_com_estoque = Estoque.objects.values('local').distinct().count()
    total_produtos = Estoque.objects.count()
    
    # Buscar endereços sem estoque
    enderecos_com_estoque_ids = Estoque.objects.values_list('local_id', flat=True).distinct()
    enderecos_sem_estoque = Armazenamento.objects.exclude(
        id__in=enderecos_com_estoque_ids
    ).order_by('rua', 'predio', 'nivel', 'ap')

    # Buscar produtos sem endereço (sem estoque)
    produtos_com_estoque_ids = Estoque.objects.values_list('produto_id', flat=True).distinct()
    produtos_sem_endereco = Produto.objects.exclude(
        id__in=produtos_com_estoque_ids
    ).order_by('nome')

    context = {
        'organizacao': organizacao_ordenada,
        'total_enderecos': total_enderecos,
        'enderecos_com_estoque': enderecos_com_estoque,
        'enderecos_vazios': total_enderecos - enderecos_com_estoque,
        'enderecos_sem_estoque': enderecos_sem_estoque,
        'produtos_sem_endereco': produtos_sem_endereco,
        'total_produtos': total_produtos,
        'taxa_ocupacao': round((enderecos_com_estoque/total_enderecos)*100, 1) if total_enderecos > 0 else 0,
        'busca_codigo': busca_codigo,
        'resultado_busca': resultado_busca
    }
    
    return render(request, 'produtos/pagina_principal.html', context)

@login_required
def detalhes_produto(request, produto_id):
    """
    Exibe todos os detalhes de um produto específico, incluindo:
    - Informações básicas do produto
    - Localização de armazenamento com botões de ação
    - Todos os lotes com validades organizados por FIFO
    - Histórico de movimentações
    - Botões de ação otimizados
    """
    produto = get_object_or_404(Produto, id=produto_id)
    
    # Buscar estoque atual organizados por níveis (FIFO: 2 -> 0)
    estoques = Estoque.objects.filter(produto=produto).select_related('local').order_by('local__nivel', 'data_armazenado')
    
    # Separar estoques por nível para lógica FIFO
    estoques_nivel_2 = estoques.filter(local__nivel='2').order_by('data_armazenado')  # Mais antigo primeiro
    estoques_nivel_0 = estoques.filter(local__nivel='0').order_by('data_armazenado')  # Mais antigo primeiro
    
    # Organizar estoques com informações extras
    estoques_processados = []
    
    # Processar nível 2 primeiro (FIFO)
    for estoque in estoques_nivel_2:
        estoque_info = {
            'estoque': estoque,
            'nivel': '2',
            'tipo': 'Armazenamento',
            'pode_abastecer': True,  # Pode abastecer para nível 0
            'nivel_destino': '0'
        }
        estoques_processados.append(estoque_info)
    
    # Processar nível 0 depois
    for estoque in estoques_nivel_0:
        estoque_info = {
            'estoque': estoque,
            'nivel': '0',
            'tipo': 'Expedição',
            'pode_abastecer': False,  # Já está no nível de saída
            'nivel_destino': None
        }
        estoques_processados.append(estoque_info)
    
    # Buscar todos os lotes ordenados por FIFO (mais antigo primeiro)
    lotes = produto.lotes.all().order_by('validade', 'created_at')
    
    # Calcular status das validades
    from datetime import date
    hoje = date.today()
    
    lotes_com_status = []
    for lote in lotes:
        if lote.validade:
            dias_para_vencer = (lote.validade - hoje).days
            
            if dias_para_vencer < 0:
                status = "Vencido"
                status_class = "status-vencido"
                prioridade = 1  # Maior prioridade
            elif dias_para_vencer <= 7:
                status = "Vence em breve"
                status_class = "status-vence-breve"
                prioridade = 2
            elif dias_para_vencer <= 30:
                status = "Próximo ao vencimento"
                status_class = "status-proximo"
                prioridade = 3
            else:
                status = "Válido"
                status_class = "status-valido"
                prioridade = 4  # Menor prioridade
        else:
            status = "Sem validade"
            status_class = "status-sem-validade"
            dias_para_vencer = None
            prioridade = 5
        
        # Verificar se este lote tem produtos no estoque de nível 2
        tem_no_nivel_2 = estoques_nivel_2.exists()
        pode_abastecer = tem_no_nivel_2 and estoques_nivel_0.count() == 0  # Só pode abastecer se não tem no nível 0
        
        # Buscar endereços onde este lote está armazenado
        # Vamos buscar por produto e data de validade, ou todos os estoques se não tiver validade específica
        if lote.validade:
            estoques_do_lote = estoques.filter(data_validade=lote.validade)
        else:
            # Se o lote não tem validade, buscar estoques sem validade
            estoques_do_lote = estoques.filter(data_validade__isnull=True)
        
        enderecos_do_lote = []
        for estoque in estoques_do_lote:
            enderecos_do_lote.append({
                'estoque': estoque,
                'endereco': estoque.local,
                'quantidade': 1  # Assumindo 1 por padrão, ajustar se necessário
            })
        
        # Se não encontrou endereços específicos, vamos pegar todos os estoques do produto
        if not enderecos_do_lote:
            for estoque in estoques[:3]:  # Limitar a 3 para não sobrecarregar
                enderecos_do_lote.append({
                    'estoque': estoque,
                    'endereco': estoque.local,
                    'quantidade': 1
                })
        
        lotes_com_status.append({
            'lote': lote,
            'status': status,
            'status_class': status_class,
            'dias_para_vencer': dias_para_vencer,
            'prioridade': prioridade,
            'pode_abastecer': pode_abastecer,
            'tem_no_nivel_2': tem_no_nivel_2,
            'enderecos': enderecos_do_lote
        })
    
    # Ordenar lotes por prioridade FIFO (vencido/próximo ao vencimento primeiro)
    lotes_com_status.sort(key=lambda x: (x['prioridade'], x['lote'].validade if x['lote'].validade else date.max))
    
    # Buscar histórico (se existir)
    try:
        from .models import HistoricoMovimentacao
        historico = HistoricoMovimentacao.objects.filter(produto=produto).order_by('-data_movimentacao')[:10]
    except:
        historico = []
    
    # Endereços disponíveis para abastecimento (nível 0)
    enderecos_nivel_0 = Armazenamento.objects.filter(nivel='0').order_by('rua', 'predio', 'ap')
    
    context = {
        'produto': produto,
        'estoques_processados': estoques_processados,
        'lotes_com_status': lotes_com_status,
        'total_lotes': lotes.count(),
        'historico': historico,
        'tem_estoque': estoques.exists(),
        'tem_nivel_2': estoques_nivel_2.exists(),
        'tem_nivel_0': estoques_nivel_0.exists(),
        'enderecos_nivel_0': enderecos_nivel_0,
        'total_enderecos': estoques.count()
    }
    
    return render(request, 'produtos/detalhes_produto.html', context)

# === SISTEMA DE ENDEREÇAMENTO MELHORADO ===

@login_required
def busca_endereco_avancada(request):
    """Busca avançada de endereços com filtros múltiplos"""
    try:
        enderecos = Armazenamento.objects.prefetch_related('estoque_set__produto').all()
        
        # Aplicar filtros
        codigo = request.GET.get('codigo', '').strip()
        codigo_produto = request.GET.get('codigo_produto', '').strip()
        rua = request.GET.get('rua', '').strip()
        predio = request.GET.get('predio', '').strip()
        nivel = request.GET.get('nivel', '').strip()
        ap = request.GET.get('ap', '').strip()
        status = request.GET.get('status', '').strip()
        tipo = request.GET.get('tipo', '').strip()
        sort = request.GET.get('sort', 'codigo')
        
        if codigo:
            enderecos = enderecos.filter(codigo__icontains=codigo)
        if codigo_produto:
            enderecos = enderecos.filter(estoque__produto__codigo__icontains=codigo_produto).distinct()
        if rua:
            enderecos = enderecos.filter(rua__icontains=rua)
        if predio:
            enderecos = enderecos.filter(predio__icontains=predio)
        if nivel:
            try:
                nivel_int = int(nivel)
                enderecos = enderecos.filter(nivel=nivel_int)
            except ValueError:
                enderecos = enderecos.filter(nivel__icontains=nivel)
        if ap:
            enderecos = enderecos.filter(ap__icontains=ap)
        if tipo:
            enderecos = enderecos.filter(categoria_armazenamento=tipo)
        
        # Converter QuerySet para lista antes dos filtros de status
        enderecos_lista = list(enderecos)
        
        # Filtro por status (vazio/ocupado)
        if status == 'vazio':
            enderecos_lista = [e for e in enderecos_lista if getattr(e, 'ocupacao_atual', lambda: 0)() == 0]
        elif status == 'ocupado':
            enderecos_lista = [e for e in enderecos_lista if getattr(e, 'ocupacao_atual', lambda: 0)() > 0]
        
        # Ordenação
        try:
            if sort == 'codigo':
                enderecos_lista.sort(key=lambda x: x.codigo or 'ZZZZ')
            elif sort == 'rua':
                enderecos_lista.sort(key=lambda x: (
                    int(x.rua) if x.rua and x.rua.isdigit() else (x.rua or 'ZZZ'), 
                    int(x.predio) if x.predio and x.predio.isdigit() else (x.predio or 'ZZZ')
                ))
            elif sort == 'ocupacao':
                enderecos_lista.sort(key=lambda x: getattr(x, 'ocupacao_atual', lambda: 0)(), reverse=True)
        except Exception as e:
            # Se houver erro na ordenação, mantém ordem original
            print(f"Erro na ordenação: {e}")
        
        return render(request, 'produtos/busca_endereco_avancada.html', {
            'enderecos': enderecos_lista,
            'total_enderecos': len(enderecos_lista),
            'filtros_aplicados': {
                'codigo': codigo,
                'codigo_produto': codigo_produto,
                'rua': rua,
                'predio': predio,
                'nivel': nivel,
                'ap': ap,
                'status': status,
                'tipo': tipo,
                'sort': sort,
            }
        })
    except Exception as e:
        # Em caso de erro, retorna página com mensagem de erro
        from django.contrib import messages
        messages.error(request, f'Erro na busca avançada: {str(e)}')
        return render(request, 'produtos/busca_endereco_avancada.html', {
            'enderecos': [],
            'total_enderecos': 0,
            'erro': str(e)
        })

@login_required
def qr_endereco(request, endereco_id):
    """Gera QR Code para um endereço específico"""
    endereco = get_object_or_404(Armazenamento, id=endereco_id)
    return render(request, 'produtos/qr_endereco.html', {
        'endereco': endereco,
    })

@login_required
def gerar_codigos_endereco(request):
    """Gera códigos únicos para todos os endereços"""
    if request.method == 'POST':
        from django.db import models
        enderecos_sem_codigo = Armazenamento.objects.filter(
            models.Q(codigo__isnull=True) | models.Q(codigo='')
        )
        
        atualizados = 0
        for endereco in enderecos_sem_codigo:
            codigo_novo = f"{str(endereco.rua).zfill(2)}-{str(endereco.predio).zfill(2)}-{str(endereco.nivel).zfill(2)}-{str(endereco.ap).zfill(2)}"
            endereco.codigo = codigo_novo
            endereco.save()
            atualizados += 1
        
        messages.success(request, f'{atualizados} endereços atualizados com códigos únicos!')
        return redirect('cadastrar_enderecos')
    
    from django.db import models
    enderecos_sem_codigo = Armazenamento.objects.filter(
        models.Q(codigo__isnull=True) | models.Q(codigo='')
    ).count()
    
    return render(request, 'produtos/gerar_codigos.html', {
        'enderecos_sem_codigo': enderecos_sem_codigo,
    })

from django.http import JsonResponse

def buscar_produto_api(request):
    """
    API para busca de produtos via AJAX - retorna dados em JSON
    """
    codigo = request.GET.get('codigo', '').strip()
    
    if not codigo:
        return JsonResponse({'success': False, 'message': 'Código não fornecido'})
    
    try:
        produto = Produto.objects.get(codigo=codigo)
        
        # Obter informações de validade
        lotes = produto.lotes.all().order_by('validade')
        proxima_validade = None
        status_validade = "Sem lote"
        dias_para_vencer = None
        
        if lotes.exists():
            primeiro_lote = lotes.first()
            proxima_validade = primeiro_lote.validade
            
            from datetime import date
            hoje = date.today()
            dias_para_vencer = (proxima_validade - hoje).days
            
            if dias_para_vencer < 0:
                status_validade = "Vencido"
            elif dias_para_vencer <= 7:
                status_validade = "Vence em breve"
            elif dias_para_vencer <= 30:
                status_validade = "Próximo ao vencimento"
            else:
                status_validade = "Válido"
        
        # Obter endereços e quantidades
        estoques = Estoque.objects.filter(produto=produto).select_related('local')
        enderecos_dict = {}
        quantidade_total = 0
        
        # Agrupar por local e contar quantidade
        for estoque in estoques:
            local_codigo = estoque.local.codigo or f"R{estoque.local.rua}P{estoque.local.predio}N{estoque.local.nivel}AP{estoque.local.ap}"
            
            if local_codigo not in enderecos_dict:
                enderecos_dict[local_codigo] = {
                    'codigo': local_codigo,
                    'tipo': estoque.local.categoria_armazenamento,
                    'quantidade': 0,
                    'rua': estoque.local.rua,
                    'predio': estoque.local.predio,
                    'nivel': estoque.local.nivel,
                    'ap': estoque.local.ap
                }
            
            enderecos_dict[local_codigo]['quantidade'] += 1
            quantidade_total += 1
        
        enderecos = list(enderecos_dict.values())
        
        return JsonResponse({
            'success': True,
            'produto': {
                'id': produto.id,
                'nome': produto.nome,
                'codigo': produto.codigo,
                'categoria': produto.categoria,
                'proxima_validade': proxima_validade.strftime('%d/%m/%Y') if proxima_validade else None,
                'status_validade': status_validade,
                'dias_para_vencer': dias_para_vencer,
                'quantidade_total': quantidade_total,
                'enderecos': enderecos
            }
        })
        
    except Produto.DoesNotExist:
        return JsonResponse({
            'success': False, 
            'message': f'Produto com código "{codigo}" não encontrado'
        })

def movimentacao_estoque(request):
    """
    Página de movimentação e abastecimento de estoque - OTIMIZADA
    """
    # Importar utils otimizados
    from .utils import ValidadeManager
    from django.db.models import Prefetch
    
    if request.method == 'POST':
        acao = request.POST.get('acao')
        codigo = request.POST.get('codigo', '').strip()
        
        if acao == 'buscar_produto':
            try:
                # Busca otimizada com prefetch
                produto = Produto.objects.prefetch_related(
                    Prefetch('lotes', queryset=Lote.objects.order_by('validade')),
                    Prefetch('estoque_set', 
                        queryset=Estoque.objects.select_related('local'))
                ).get(codigo=codigo)
                
                # Usar listas ao invés de querysets para performance
                estoques = list(produto.estoque_set.all())
                lotes = list(produto.lotes.all())
                todas_validades = [lote.validade for lote in lotes if lote.validade]
                
                # Data atual
                data_atual = date.today()
                
                # Verificar se o produto não tem estoque
                if not estoques:
                    context = {
                        'produto_encontrado': produto,
                        'produto_sem_estoque': True,
                        'enderecos_info': [],
                        'total_unidades': 0,
                        'codigo_busca': codigo,
                        'data_atual': data_atual,
                        'todas_validades': todas_validades
                    }
                else:
                    # Usar função otimizada para distribuir validades
                    enderecos_info = ValidadeManager.distribuir_validades_por_cards(
                        estoques, todas_validades
                    )
                    
                    context = {
                        'produto_encontrado': produto,
                        'enderecos_info': enderecos_info,
                        'total_unidades': len(estoques),
                        'codigo_busca': codigo,
                        'data_atual': data_atual,
                        'proxima_validade': todas_validades[0] if todas_validades else None,
                        'todas_validades': todas_validades
                    }
                
            except Produto.DoesNotExist:
                messages.error(request, f'Produto com código "{codigo}" não encontrado.')
                context = {}
            except Exception as e:
                messages.error(request, f'Erro ao buscar produto: {str(e)}')
                context = {}
                
        elif acao == 'alocar_produto':
            produto_id = request.POST.get('produto_id')
            endereco_destino_id = request.POST.get('endereco_destino_id')
            data_validade = request.POST.get('data_validade')
            
            try:
                produto = Produto.objects.get(id=produto_id)
                endereco_destino = Armazenamento.objects.get(id=endereco_destino_id)
                
                # Criar novo estoque
                novo_estoque = Estoque.objects.create(
                    produto=produto,
                    local=endereco_destino,
                    data_armazenado=timezone.now().date(),
                    data_validade=data_validade if data_validade else None,
                    data_alteracao=timezone.now(),
                    usuario_responsavel=request.user.username if hasattr(request.user, 'username') else 'Sistema'
                )
                
                # Registrar movimentação
                HistoricoMovimentacao.objects.create(
                    produto=produto,
                    local_origem=None,
                    local_destino=endereco_destino,
                    tipo_operacao='entrada',
                    quantidade=1,
                    usuario=request.user.username if hasattr(request.user, 'username') else 'Sistema',
                    observacoes=f'Alocação inicial do produto - Validade: {data_validade if data_validade else "Não informada"}'
                )
                
                messages.success(request, f'Produto alocado em {endereco_destino.codigo or endereco_destino} com sucesso!')
                context = {}
                
            except (Produto.DoesNotExist, Armazenamento.DoesNotExist):
                messages.error(request, 'Erro ao alocar produto.')
                context = {}
            
        elif acao == 'abastecer':
            estoque_origem_id = request.POST.get('estoque_origem_id')
            endereco_destino_id = request.POST.get('endereco_destino_id')
            
            try:
                estoque_origem = Estoque.objects.get(id=estoque_origem_id)
                endereco_destino = Armazenamento.objects.get(id=endereco_destino_id)
                produto = estoque_origem.produto
                
                # Verificar se o endereço de destino está no nível 0 (área de saída)
                if endereco_destino.nivel != '0':
                    messages.error(request, 'Abastecimento só pode ser feito para endereços no nível 0 (área de saída).')
                    return redirect('movimentacao_estoque')
                
                # Verificar se já existe produto no nível 0 (destino)
                estoque_destino_existente = Estoque.objects.filter(
                    produto=produto,
                    local=endereco_destino
                ).first()
                
                if estoque_destino_existente:
                    # Produto já existe no nível 0 - atualizar com data mais antiga (FIFO)
                    if estoque_origem.data_armazenado < estoque_destino_existente.data_armazenado or (estoque_origem.data_validade and estoque_destino_existente.data_validade and estoque_origem.data_validade < estoque_destino_existente.data_validade):
                        # Produto do nível 2 é mais antigo ou tem validade mais próxima, atualizar o nível 0
                        data_anterior = estoque_destino_existente.data_armazenado
                        validade_anterior = estoque_destino_existente.data_validade
                        
                        estoque_destino_existente.data_armazenado = estoque_origem.data_armazenado
                        estoque_destino_existente.data_validade = estoque_origem.data_validade
                        estoque_destino_existente.data_alteracao = timezone.now()
                        estoque_destino_existente.save()
                        
                        # Remover produto do nível 2
                        local_origem = estoque_origem.local
                        estoque_origem.delete()
                        
                        # Registrar movimentação
                        HistoricoMovimentacao.objects.create(
                            produto=produto,
                            local_origem=local_origem,
                            local_destino=endereco_destino,
                            tipo_operacao='atualizacao_fifo',
                            quantidade=1,
                            usuario=request.user.username if hasattr(request.user, 'username') else 'Sistema',
                            observacoes=f'Atualização FIFO - data alterada de {data_anterior.strftime("%d/%m/%Y")} para {estoque_origem.data_armazenado.strftime("%d/%m/%Y")}' + (f', validade de {validade_anterior.strftime("%d/%m/%Y") if validade_anterior else "sem validade"} para {estoque_origem.data_validade.strftime("%d/%m/%Y") if estoque_origem.data_validade else "sem validade"}' if estoque_origem.data_validade != validade_anterior else '')
                        )
                        
                        messages.success(request, f'Produto no nível 0 atualizado com data mais antiga ({estoque_origem.data_armazenado.strftime("%d/%m/%Y")}) e validade atualizada!')
                        
                        # Verificar se o nível 2 ficou vazio
                        if not Estoque.objects.filter(local=local_origem).exists():
                            context = {
                                'pergunta_excluir_nivel2': True,
                                'endereco_nivel2': local_origem,
                                'produto_movido': produto
                            }
                        else:
                            context = {}
                    else:
                        # Produto no nível 0 já é mais antigo
                        messages.info(request, f'Produto no nível 0 já possui data/validade mais antiga. Nenhuma alteração necessária.')
                        context = {}
                else:
                    # Não existe produto no nível 0, mover do nível 2 para nível 0
                    local_origem = estoque_origem.local
                    estoque_origem.local = endereco_destino
                    estoque_origem.data_alteracao = timezone.now()
                    estoque_origem.save()
                    
                    # Registrar movimentação
                    HistoricoMovimentacao.objects.create(
                        produto=produto,
                        local_origem=local_origem,
                        local_destino=endereco_destino,
                        tipo_operacao='transferencia',
                        quantidade=1,
                        usuario=request.user.username if hasattr(request.user, 'username') else 'Sistema',
                        observacoes=f'Movimentação do nível 2 para nível 0 (área de saída) - Validade: {estoque_origem.data_validade.strftime("%d/%m/%Y") if estoque_origem.data_validade else "Não informada"}'
                    )
                    
                    messages.success(request, f'Produto movido para {endereco_destino.codigo or endereco_destino} com sucesso!')
                    
                    # Verificar se o nível 2 ficou vazio
                    if not Estoque.objects.filter(local=local_origem).exists():
                        context = {
                            'pergunta_excluir_nivel2': True,
                            'endereco_nivel2': local_origem,
                            'produto_movido': produto
                        }
                    else:
                        context = {}
                
            except (Estoque.DoesNotExist, Armazenamento.DoesNotExist):
                messages.error(request, 'Erro ao processar abastecimento.')
                context = {}
                
        elif acao == 'excluir_nivel2':
            endereco_id = request.POST.get('endereco_id')
            try:
                endereco = Armazenamento.objects.get(id=endereco_id)
                endereco.delete()
                messages.success(request, f'Endereço {endereco} excluído com sucesso.')
            except Armazenamento.DoesNotExist:
                messages.error(request, 'Endereço não encontrado.')
            context = {}
            
        elif acao == 'manter_nivel2':
            messages.info(request, 'Endereço mantido no nível 2.')
            context = {}
                
        else:
            context = {}
    else:
        context = {}
    
    # Buscar endereços disponíveis para abastecimento (nível 0)
    enderecos_destino = Armazenamento.objects.filter(nivel='0').order_by('rua', 'predio', 'ap')
    
    context.update({
        'enderecos_destino': enderecos_destino,
        'historico_recente': HistoricoMovimentacao.objects.filter(
            tipo_operacao='transferencia'
        ).order_by('-data_operacao')[:10]
    })
    
    return render(request, 'produtos/movimentacao_estoque.html', context)

from django.views.decorators.csrf import csrf_exempt

@login_required
@csrf_exempt
def conferente_rapido(request):
    """Modo Conferente Rápido: busca instantânea por código"""
    produto = None
    proxima_validade = None
    localizacao = None
    status_urgencia = None
    codigo_busca = None
    if request.method == 'POST':
        codigo_busca = request.POST.get('codigo', '').strip()
        try:
            produto = Produto.objects.get(codigo=codigo_busca)
            # Buscar estoque mais recente
            estoque = Estoque.objects.filter(produto=produto).order_by('data_armazenado').first()
            localizacao = estoque.local.codigo if estoque and estoque.local else '---'
            # Buscar próxima validade
            lote = produto.lotes.filter(validade__isnull=False).order_by('validade').first()
            proxima_validade = lote.validade if lote else None
            # Calcular status de urgência
            if proxima_validade:
                hoje = date.today()
                dias = (proxima_validade - hoje).days
                if dias < 0:
                    status_urgencia = 'VENCIDO'
                elif dias <= 7:
                    status_urgencia = 'VENCE_EM_BREVE'
                else:
                    status_urgencia = 'OK'
            else:
                status_urgencia = 'SEM_VALIDADE'
        except Produto.DoesNotExist:
            produto = None
    return render(request, 'produtos/conferente_rapido.html', {
        'produto': produto,
        'proxima_validade': proxima_validade,
        'localizacao': localizacao,
        'status_urgencia': status_urgencia,
        'codigo_busca': codigo_busca
    })